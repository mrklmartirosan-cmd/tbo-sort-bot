import os
import json
import base64
import logging
import re
import io
import asyncio
from datetime import datetime, timezone, timedelta
from telegram import Update, ReplyKeyboardMarkup, InputFile
from telegram.ext import (
    Application, CommandHandler, MessageHandler, filters,
    ContextTypes, ConversationHandler
)
import httpx
import gspread
import openpyxl

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.environ.get("BOT_TOKEN", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
WEBHOOK_URL = os.environ.get("WEBHOOK_URL", "")
PORT = int(os.environ.get("PORT", 8080))

# --- Google Sheets ---
SPREADSHEET_ID = os.environ.get("SPREADSHEET_ID", "")
REPORT_SPREADSHEET_ID = os.environ.get("REPORT_SPREADSHEET_ID", "")  # файл "Реализация 2026г"
GOOGLE_CREDENTIALS = os.environ.get("GOOGLE_CREDENTIALS", "")  # содержимое JSON-ключа целиком

SHEET_PRODUCTION = "Производство"
SHEET_SALES = "Реализация"

# Список разрешённых Telegram ID (закрытый доступ).
# Пустой список = пускать всех (страховка, чтобы не заблокировать себя).
_allowed_raw = os.environ.get("ALLOWED_USERS", "")
ALLOWED_USERS = {int(x) for x in _allowed_raw.replace(" ", "").split(",") if x.strip().isdigit()}

# --- Автобэкап ---
BACKUP_CHANNEL_ID = os.environ.get("BACKUP_CHANNEL_ID", "")
KZ_TZ = timezone(timedelta(hours=5))
BACKUP_HOUR = 22
_last_backup_counts = {}

# --- настройки отчёта реализации ---------------------------------
REPORT_TEMPLATE = "ШАБЛОН"                 # лист-образец (пустой бланк)
REPORT_DATE_COL = 2                        # B = Дата
REPORT_BUYER_COL = 3                       # C = Контрагент
# Для каждой фракции: (Кол-во, Цена, Сумма) — номера столбцов (1-based)
REPORT_FRAC_COLS = {
    "0-1": (4, 5, 6),     # D E F
    "1-2": (7, 8, 9),     # G H I
    "2-4": (10, 11, 12),  # J K L
    "4-6": (13, 14, 15),  # M N O
    "6-8": (16, 17, 18),  # P Q R
}
# Блоки (строки данных, без строки "Итого")
REPORT_BLOCKS = {
    "Безналичный": (8, 23),
    "Наличный": (26, 40),
}
RU_MONTHS = {
    1: "январь", 2: "февраль", 3: "март", 4: "апрель", 5: "май", 6: "июнь",
    7: "июль", 8: "август", 9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь",
}
# ------------------------------------------------------------------

# Состояния диалогов ручного ввода
(P_DATE, P_FIO, P_TIRES, P_BAGS, P_THREAD, P_F01, P_F12, P_F24, P_F46, P_F68, P_CORD, P_NOTE) = range(12)
(S_DATE, S_BUYER, S_PAYTYPE, S_FRAC, S_KG, S_PRICE, S_SUM_VAT, S_VAT, S_NOTE) = range(12, 21)
# NEW: состояние подтверждения дубля при фото-производстве
# + доспрос нал/безнал после фото-реализации (чтобы продажа попала в отчёт)
(PHOTO_PROD_CONFIRM, PHOTO_SALE_PAY, PHOTO_TYPE_CONFIRM) = range(21, 24)
(E_DATE, E_AMOUNT, E_GROUP, E_CATEGORY, E_SOURCE, E_CONTRAGENT, E_NOTE, E_CONFIRM, E_SUP_CONFIRM, E_SUP_NEW, E_SUP_EDIT) = range(24, 35)
# NEW: банк поступления при безнал-продаже (ручной ввод и фото) — для финотчёта
(S_BANK, PHOTO_SALE_BANK) = range(35, 37)
# NEW: доспрос недостающего по фото-продаже (платёжка без кол-ва/цены/НДС)
(PHOTO_SALE_KG, PHOTO_SALE_PRICE, PHOTO_SALE_VAT) = range(37, 40)
# NEW: ведомость ФОТ (зарплата): дата перечисления, банк, подтверждение
(PAYROLL_DATE, PAYROLL_BANK, PAYROLL_CONFIRM) = range(40, 43)
# NEW: табель учёта рабочего времени
(TABEL_MONTH, TABEL_CONFIRM) = range(43, 45)

FRACTIONS = ["0-1", "1-2", "2-4", "4-6", "6-8"]

# Заголовки листов склада (порядок колонок = порядок записи)
PROD_HEADERS = ["Дата", "ФИО оператора", "Вес шин кг", "Мешки шт", "Нитки",
                "Фракция 0-1", "Фракция 1-2", "Фракция 2-4", "Фракция 4-6", "Фракция 6-8",
                "Всего крошки кг", "Металлокорд кг", "Примечание"]
SALES_HEADERS = ["Дата", "Покупатель", "Тип расчёта", "Фракция", "Количество кг",
                 "Цена за кг", "Сумма с НДС", "Сумма НДС", "Примечание"]

_gc = None
_spreadsheet = None
_report_spreadsheet = None


def parse_num(text):
    """Превращает '210', '210,5', ' 210 кг ' в число. Пустое/прочерк -> 0."""
    if text is None:
        return 0
    t = str(text).strip().lower().replace(",", ".")
    t = t.replace("кг", "").replace("тнг", "").replace(" ", "")
    if t in ("", "-", "—", "нет"):
        return 0
    try:
        num = float(re.sub(r"[^0-9.]", "", t))
        return int(num) if num == int(num) else num
    except Exception:
        return 0


def parse_date_or_none(date_text):
    """Главная проверка даты. Принимает строку, которую отдал Claude или ввёл человек.
    Понимает обычные форматы С РАЗДЕЛИТЕЛЯМИ: '08.06.2026', '8.6.26', '8/6/2026', '8-6-26'.
    Год из двух цифр -> 20xx. Возвращает datetime, либо None если разобрать нельзя.
    НИКАКОГО угадывания слитных цифр — если непонятно, вернём None и спросим человека."""
    if date_text is None:
        return None
    s = str(date_text).strip()
    if not s:
        return None
    # сперва пробуем строгие форматы целиком
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    # затем — разделители любого вида: ровно 3 числовые части день/месяц/год
    parts = [p for p in re.split(r"[.\-/\s]+", s) if p != ""]
    if len(parts) == 3 and all(p.isdigit() for p in parts):
        d, m, y = parts
        try:
            d_i, m_i, y_i = int(d), int(m), int(y)
            if y_i < 100:
                y_i += 2000
            return datetime(y_i, m_i, d_i)
        except Exception:
            return None
    return None


def date_to_str(dt):
    """datetime -> 'дд.мм.гггг'."""
    return dt.strftime("%d.%m.%Y")


# ---------- контроль доступа ----------

def is_allowed(update: Update) -> bool:
    """True, если пользователь в списке разрешённых.
    Если ALLOWED_USERS пуст — пускаем всех (страховка от самоблокировки)."""
    if not ALLOWED_USERS:
        return True
    user = update.effective_user
    return bool(user and user.id in ALLOWED_USERS)


# ---------- Google Sheets: подключение и доступ к листам ----------

def _get_client():
    global _gc
    if _gc is None:
        if not GOOGLE_CREDENTIALS:
            raise RuntimeError("Не задан GOOGLE_CREDENTIALS в переменных окружения.")
        creds_dict = json.loads(GOOGLE_CREDENTIALS)
        _gc = gspread.service_account_from_dict(creds_dict)
    return _gc


def get_spreadsheet():
    global _spreadsheet
    if _spreadsheet is not None:
        return _spreadsheet
    if not SPREADSHEET_ID:
        raise RuntimeError("Не задан SPREADSHEET_ID в переменных окружения.")
    _spreadsheet = _get_client().open_by_key(SPREADSHEET_ID)
    return _spreadsheet


def get_report_spreadsheet():
    global _report_spreadsheet
    if _report_spreadsheet is not None:
        return _report_spreadsheet
    if not REPORT_SPREADSHEET_ID:
        raise RuntimeError("Не задан REPORT_SPREADSHEET_ID в переменных окружения.")
    _report_spreadsheet = _get_client().open_by_key(REPORT_SPREADSHEET_ID)
    return _report_spreadsheet


def get_worksheet(title, headers):
    sh = get_spreadsheet()
    try:
        ws = sh.worksheet(title)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=title, rows=1000, cols=max(len(headers), 12))
        ws.append_row(headers, value_input_option="USER_ENTERED")
        return ws
    first = ws.row_values(1)
    if not first:
        ws.append_row(headers, value_input_option="USER_ENTERED")
    return ws


def get_production_ws():
    return get_worksheet(SHEET_PRODUCTION, PROD_HEADERS)


def get_sales_ws():
    return get_worksheet(SHEET_SALES, SALES_HEADERS)


# ---------- Чтение/запись данных ----------

def _fractions_signature(data):
    """NEW: подпись записи по фракциям (для сравнения дублей).
    Кортеж из 5 чисел фракций — устойчив к разнице в ФИО/примечании."""
    return (
        parse_num(data.get("фракция_0_1", 0)),
        parse_num(data.get("фракция_1_2", 0)),
        parse_num(data.get("фракция_2_4", 0)),
        parse_num(data.get("фракция_4_6", 0)),
        parse_num(data.get("фракция_6_8", 0)),
    )


def find_duplicate_production(data):
    """NEW: ищет в складе строку с той же датой И теми же фракциями.
    Возвращает True, если такая уже есть (вероятный дубль того же бланка).
    Разные смены за один день с разными цифрами дублем НЕ считаются."""
    ws = get_production_ws()
    rows = ws.get_all_values()
    if len(rows) <= 1:
        return False

    target_dt = parse_date_or_none(data.get("дата", ""))
    target_sig = _fractions_signature(data)

    for r in rows[1:]:
        row_date = r[0] if len(r) > 0 else ""
        # сравниваем даты по смыслу (а не по тексту): 08.06.2026 == 8.6.26
        row_dt = parse_date_or_none(row_date)
        same_date = False
        if target_dt and row_dt:
            same_date = (target_dt.date() == row_dt.date())
        else:
            same_date = (str(row_date).strip() == str(data.get("дата", "")).strip())
        if not same_date:
            continue
        # сравниваем фракции (индексы 5..9 в PROD_HEADERS)
        row_sig = tuple(
            parse_num(r[i]) if i < len(r) else 0
            for i in range(5, 10)
        )
        if row_sig == target_sig:
            return True
    return False


def save_production(data):
    ws = get_production_ws()
    f01 = parse_num(data.get("фракция_0_1", 0))
    f12 = parse_num(data.get("фракция_1_2", 0))
    f24 = parse_num(data.get("фракция_2_4", 0))
    f46 = parse_num(data.get("фракция_4_6", 0))
    f68 = parse_num(data.get("фракция_6_8", 0))
    total = f01 + f12 + f24 + f46 + f68
    bags = round(total / 30) if total else 0  # 1 мешок = 30 кг; считаем сами, не из распознавания
    row = [
        data.get("дата", datetime.now().strftime("%d.%m.%Y")),
        data.get("фио", ""),
        parse_num(data.get("вес_шин", 0)),
        bags,
        parse_num(data.get("нитки", 0)),
        f01, f12, f24, f46, f68,
        total,
        parse_num(data.get("металл_корд", 0)),
        data.get("примечание", ""),
    ]
    ws.append_row(row, value_input_option="USER_ENTERED")


def _note_with_bank(data):
    """Примечание продажи + банк поступления (нужно финотчёту; колонки листа не меняем)."""
    note = data.get("примечание", "")
    bank = data.get("банк", "")
    if bank:
        note = (note + "; " if note else "") + f"банк: {bank}"
    return note


def save_sale(data):
    ws = get_sales_ws()
    qty_kg = parse_num(data.get("количество_кг", 0))
    if not qty_kg and data.get("количество_т"):
        qty_kg = parse_num(data.get("количество_т", 0)) * 1000
    row = [
        data.get("дата", datetime.now().strftime("%d.%m.%Y")),
        data.get("покупатель", ""),
        data.get("тип_расчета", ""),
        data.get("фракция", ""),
        qty_kg,
        parse_num(data.get("цена_за_кг", 0)),
        parse_num(data.get("сумма_с_ндс", 0)),
        parse_num(data.get("сумма_ндс", 0)),
        _note_with_bank(data),
    ]
    ws.append_row(row, value_input_option="USER_ENTERED")


# ---------- запись реализации в отчёт "Реализация 2026г" ----------

def _normalize_fraction(frac_text):
    for f in FRACTIONS:
        if f in str(frac_text):
            return f
    return None


def _get_or_create_month_sheet(sh, dt):
    title = f"{RU_MONTHS[dt.month]} {dt.year}"
    try:
        return sh.worksheet(title)
    except gspread.WorksheetNotFound:
        pass
    template = sh.worksheet(REPORT_TEMPLATE)
    new_ws = template.duplicate(new_sheet_name=title)
    return new_ws


def write_sale_to_report(data):
    """Дописывает строку реализации в отчёт. Возвращает (ok: bool, msg: str)."""
    pay = str(data.get("тип_расчета", "")).strip()
    if pay not in REPORT_BLOCKS:
        return False, f"тип расчёта не распознан ('{pay}') — строка в отчёт не добавлена"

    frac = _normalize_fraction(data.get("фракция", ""))
    if frac is None:
        return False, "фракция не распознана — строка в отчёт не добавлена"

    dt = parse_date_or_none(data.get("дата", ""))
    if dt is None:
        return False, "дата не распознана — строка в отчёт не добавлена"

    sh = get_report_spreadsheet()
    ws = _get_or_create_month_sheet(sh, dt)

    start_row, end_row = REPORT_BLOCKS[pay]
    date_col_vals = ws.col_values(REPORT_DATE_COL)
    target_row = None
    for r in range(start_row, end_row + 1):
        val = date_col_vals[r - 1] if r - 1 < len(date_col_vals) else ""
        if not str(val).strip():
            target_row = r
            break
    if target_row is None:
        return False, f"блок «{pay}» заполнен (нет свободных строк) — строка не добавлена"

    qty_kg = parse_num(data.get("количество_кг", 0))
    if not qty_kg and data.get("количество_т"):
        qty_kg = parse_num(data.get("количество_т", 0)) * 1000
    price = parse_num(data.get("цена_за_кг", 0))
    total_sum = parse_num(data.get("сумма_с_ндс", 0))

    qcol, pcol, scol = REPORT_FRAC_COLS[frac]

    updates = [
        {"range": gspread.utils.rowcol_to_a1(target_row, REPORT_DATE_COL),
         "values": [[dt.strftime("%d.%m.%Y")]]},
        {"range": gspread.utils.rowcol_to_a1(target_row, REPORT_BUYER_COL),
         "values": [[data.get("покупатель", "")]]},
        {"range": gspread.utils.rowcol_to_a1(target_row, qcol), "values": [[qty_kg]]},
        {"range": gspread.utils.rowcol_to_a1(target_row, pcol), "values": [[price]]},
        {"range": gspread.utils.rowcol_to_a1(target_row, scol), "values": [[total_sum]]},
    ]
    ws.batch_update(updates, value_input_option="USER_ENTERED")
    return True, f"добавлено в «{ws.title}», блок «{pay}», строка {target_row}"


# ---------- Остаток ----------

def calc_stock():
    income = {k: 0 for k in FRACTIONS}
    ws_p = get_production_ws()
    rows = ws_p.get_all_values()
    for r in rows[1:]:
        for i, key in enumerate(FRACTIONS, 5):
            if i < len(r):
                try:
                    income[key] += float(str(r[i]).replace(",", ".")) if r[i] else 0
                except Exception:
                    pass

    outcome = {k: 0 for k in FRACTIONS}
    ws_s = get_sales_ws()
    rows_s = ws_s.get_all_values()
    for r in rows_s[1:]:
        frac = r[3] if len(r) > 3 else ""
        qty = r[4] if len(r) > 4 else 0
        for key in outcome:
            if key in str(frac):
                try:
                    outcome[key] += float(str(qty).replace(",", ".")) if qty else 0
                except Exception:
                    pass
                break

    stock = {k: income[k] - outcome[k] for k in income}
    return income, outcome, stock


# ---------- Распознавание фото через Claude ----------

async def call_claude(image_b64: str, prompt: str):
    async with httpx.AsyncClient(timeout=60) as client:
        response = await client.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json"
            },
            json={
                "model": "claude-opus-4-5",
                "max_tokens": 2000,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": image_b64}},
                        {"type": "text", "text": prompt}
                    ]
                }]
            }
        )
        result = response.json()
        logger.info(f"Claude response: {result}")
        text = result["content"][0]["text"].strip()
        text = re.sub(r'```json|```', '', text).strip()
        return json.loads(text)


async def recognize_production(image_bytes: bytes):
    image_b64 = base64.b64encode(image_bytes).decode()
    prompt = """Это фото ежедневного отчёта по производству резиновой крошки.
На фото может быть одна или несколько строк. Верни ТОЛЬКО JSON без markdown.
Если одна строка — объект, если несколько — массив объектов. Формат:
{"дата":"дд.мм.гггг","фио":"ФИО","вес_шин":0,"мешки":0,"нитки":0,"фракция_0_1":0,"фракция_1_2":0,"фракция_2_4":0,"фракция_4_6":0,"фракция_6_8":0,"всего_итог":0,"металл_корд":0,"примечание":""}

ВАЖНО про ДАТУ:
- На бланке дату часто пишут слитно, без нулей и точек: формат деньмесяцгод.
  Пример: "5626" означает 5.6.26 -> верни "05.06.2026".
  Пример: "151226" означает 15.12.26 -> верни "15.12.2026".
- Год пишут двумя цифрами (26), это 2026 -> в ответе год всегда 4 цифры (2026).
- В ответе дата ВСЕГДА в виде "дд.мм.гггг" с точками и ведущими нулями.

ВАЖНО про ФРАКЦИИ:
- В одной ячейке фракции может стоять ДВА числа: верхнее = мешки (штуки), нижнее = килограммы.
- Бери ВСЕГДА НИЖНЕЕ число (килограммы). Верхнее (мешки) игнорируй.
- Если число одно — это килограммы.

ВАЖНО про ВСЕГО:
- На бланке есть итоговое поле «всего крошки» (общий вес за смену в кг).
- Прочитай это число и верни в "всего_итог" (если два числа — бери нижнее, кг). Если поля нет — 0.

Если что-то не читается — ставь 0."""
    return await call_claude(image_b64, prompt)


async def recognize_sale(image_bytes: bytes):
    image_b64 = base64.b64encode(image_bytes).decode()
    prompt = """Это фото накладной на отпуск/реализацию резиновой крошки.
Распознай данные и верни ТОЛЬКО JSON без markdown:
{"дата":"дд.мм.гггг","покупатель":"название организации или ИП","фракция":"например 2-4","количество_т":0.0,"количество_кг":0,"цена_за_кг":0,"сумма_с_ндс":0,"сумма_ндс":0,"примечание":""}
Количество бери из колонки "отпущено". Если оно в тоннах — заполни количество_т, если в кг — количество_кг.
Цену бери из колонки "цена" (цена за 1 кг). Бери ровно как в накладной, не пересчитывай.
Сумму с НДС бери из колонки "Сумма с НДС" (полная сумма включая НДС).
Сумму НДС бери из колонки "Сумма НДС" (только НДС). Если НДС в накладной нет — поставь 0.
Дату верни в виде "дд.мм.гггг". Если не читается — 0."""
    return await call_claude(image_b64, prompt)


# ---------- Команды и меню ----------

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("⛔ Нет доступа. Обратитесь к администратору.")
        return
    keyboard = [
        ["📸 Производство", "📄 Реализация"],
        ["✍️ Ввод производства", "✍️ Ввод реализации"],
        ["💸 Ввод расхода"],
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        "👋 Привет! Я бот учёта резиновой крошки.\n\n"
        "📸 Производство — распознать фото отчёта\n"
        "📄 Реализация — распознать фото накладной\n"
        "✍️ Ввод производства — внести вручную по шагам\n"
        "✍️ Ввод реализации — внести вручную по шагам\n\n"
        "Команды:\n"
        "/ostatok — остаток на складе\n"
        "/last — последние записи\n"
        "/cancel — отменить ручной ввод",
        reply_markup=reply_markup
    )


async def ostatok(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("⛔ Нет доступа. Обратитесь к администратору.")
        return
    income, outcome, stock = calc_stock()
    total_in = sum(income.values())
    total_out = sum(outcome.values())
    total_stock = sum(stock.values())

    text = "📦 *ОСТАТОК НА СКЛАДЕ*\n\n"
    text += "```\n"
    text += f"{'Фракция':<8} {'Приход':>8} {'Расход':>8} {'Остаток':>9}\n"
    text += "-" * 37 + "\n"
    for key in FRACTIONS:
        text += f"{key:<8} {income[key]:>8.0f} {outcome[key]:>8.0f} {stock[key]:>9.0f}\n"
    text += "-" * 37 + "\n"
    text += f"{'ИТОГО':<8} {total_in:>8.0f} {total_out:>8.0f} {total_stock:>9.0f}\n"
    text += "```\n_Все данные в кг_"

    await update.message.reply_text(text, parse_mode="Markdown")


async def last_records(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("⛔ Нет доступа. Обратитесь к администратору.")
        return
    ws = get_production_ws()
    rows = ws.get_all_values()
    data_rows = rows[1:]
    if not data_rows:
        await update.message.reply_text("📭 Записей пока нет.")
        return
    last5 = data_rows[-5:]
    text = "📋 *Последние записи производства:*\n\n"
    for r in last5:
        date = r[0] if len(r) > 0 and r[0] else "—"
        fio = r[1] if len(r) > 1 and r[1] else "—"
        total = r[10] if len(r) > 10 and r[10] else "—"
        text += f"📅 {date} | {fio} | {total} кг\n"
    await update.message.reply_text(text, parse_mode="Markdown")


# ---------- Ручной ввод: ПРОИЗВОДСТВО ----------

async def manual_prod_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("⛔ Нет доступа. Обратитесь к администратору.")
        return ConversationHandler.END
    context.user_data["prod"] = {}
    await update.message.reply_text(
        "✍️ Ручной ввод производства.\n\n"
        "Введи *дату* (дд.мм.гггг).\n"
        "Или напиши «сегодня».\n\n"
        "Для отмены — /cancel",
        parse_mode="Markdown"
    )
    return P_DATE


async def manual_prod_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    t = update.message.text.strip()
    if t.lower() == "сегодня":
        context.user_data["prod"]["дата"] = datetime.now().strftime("%d.%m.%Y")
    else:
        dt = parse_date_or_none(t)
        if dt is None:
            await update.message.reply_text(
                "⚠️ Не понял дату. Введи в виде ДД.ММ.ГГГГ (например 05.06.2026)\n"
                "или напиши «сегодня». Для отмены — /cancel"
            )
            return P_DATE
        context.user_data["prod"]["дата"] = date_to_str(dt)
    await update.message.reply_text("👤 ФИО оператора?")
    return P_FIO


async def manual_prod_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["prod"]["фио"] = update.message.text.strip()
    await update.message.reply_text("⚖️ Вес шин, кг?")
    return P_TIRES


async def manual_prod_tires(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["prod"]["вес_шин"] = parse_num(update.message.text)
    await update.message.reply_text("🧵 Нитки, кг? (или «-» если нет)")
    return P_THREAD


async def manual_prod_thread(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["prod"]["нитки"] = parse_num(update.message.text)
    await update.message.reply_text("Фракция 0-1, кг?")
    return P_F01


async def manual_prod_f01(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["prod"]["фракция_0_1"] = parse_num(update.message.text)
    await update.message.reply_text("Фракция 1-2, кг?")
    return P_F12


async def manual_prod_f12(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["prod"]["фракция_1_2"] = parse_num(update.message.text)
    await update.message.reply_text("Фракция 2-4, кг?")
    return P_F24


async def manual_prod_f24(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["prod"]["фракция_2_4"] = parse_num(update.message.text)
    await update.message.reply_text("Фракция 4-6, кг?")
    return P_F46


async def manual_prod_f46(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["prod"]["фракция_4_6"] = parse_num(update.message.text)
    await update.message.reply_text("Фракция 6-8, кг?")
    return P_F68


async def manual_prod_f68(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["prod"]["фракция_6_8"] = parse_num(update.message.text)
    await update.message.reply_text("🔩 Металлокорд, кг?")
    return P_CORD


async def manual_prod_cord(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["prod"]["металл_корд"] = parse_num(update.message.text)
    await update.message.reply_text("📝 Примечание? (или «-» если нет)")
    return P_NOTE


def _prod_summary(d):
    """Текст-подтверждение сохранённой записи производства."""
    total = (parse_num(d.get("фракция_0_1", 0)) + parse_num(d.get("фракция_1_2", 0))
             + parse_num(d.get("фракция_2_4", 0)) + parse_num(d.get("фракция_4_6", 0))
             + parse_num(d.get("фракция_6_8", 0)))
    bags = round(total / 30) if total else 0
    return (
        f"✅ *Производство сохранено!*\n\n"
        f"📅 Дата: {d.get('дата', '—')}\n"
        f"👤 Оператор: {d.get('фио', '—')}\n"
        f"⚖️ Вес шин: {d.get('вес_шин', 0)} кг\n"
        f"🛍 Мешки: {bags} шт (всего ÷ 30)\n\n"
        f"*Фракции (кг):*\n"
        f"  0-1: {d.get('фракция_0_1', 0)}\n"
        f"  1-2: {d.get('фракция_1_2', 0)}\n"
        f"  2-4: {d.get('фракция_2_4', 0)}\n"
        f"  4-6: {d.get('фракция_4_6', 0)}\n"
        f"  6-8: {d.get('фракция_6_8', 0)}\n"
        f"  Всего: {total} кг\n\n"
        f"🔩 Металлокорд: {d.get('металл_корд', 0)} кг"
    )


async def manual_prod_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    note = update.message.text.strip()
    context.user_data["prod"]["примечание"] = "" if note in ("-", "—") else note
    d = context.user_data["prod"]

    # NEW: при ручном вводе тоже предупреждаем о вероятном дубле (но сохраняем — это явное действие человека)
    dup_note = ""
    try:
        if find_duplicate_production(d):
            dup_note = "\n\n⚠️ Похоже, запись с этой датой и теми же фракциями уже есть — добавил ещё одну (ручной ввод)."
    except Exception as e:
        logger.error(f"dup check error (manual): {e}")

    save_production(d)
    schedule_refresh()
    await update.message.reply_text(_prod_summary(d) + dup_note, parse_mode="Markdown")
    context.user_data.pop("prod", None)
    return ConversationHandler.END


# ---------- Ручной ввод: РЕАЛИЗАЦИЯ ----------

async def manual_sale_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("⛔ Нет доступа. Обратитесь к администратору.")
        return ConversationHandler.END
    context.user_data["sale"] = {}
    await update.message.reply_text(
        "✍️ Ручной ввод реализации.\n\n"
        "Введи *дату* (дд.мм.гггг) или «сегодня».\n\n"
        "Для отмены — /cancel",
        parse_mode="Markdown"
    )
    return S_DATE


async def manual_sale_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    t = update.message.text.strip()
    if t.lower() == "сегодня":
        context.user_data["sale"]["дата"] = datetime.now().strftime("%d.%m.%Y")
    else:
        dt = parse_date_or_none(t)
        if dt is None:
            await update.message.reply_text(
                "⚠️ Не понял дату. Введи в виде ДД.ММ.ГГГГ (например 05.06.2026)\n"
                "или напиши «сегодня». Для отмены — /cancel"
            )
            return S_DATE
        context.user_data["sale"]["дата"] = date_to_str(dt)
    await update.message.reply_text("🏢 Покупатель?")
    return S_BUYER


async def manual_sale_buyer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["sale"]["покупатель"] = update.message.text.strip()
    kb = ReplyKeyboardMarkup([["Безналичный", "Наличный"], ["❌ Отмена"]], resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text("💳 Тип расчёта? (Безналичный / Наличный)", reply_markup=kb)
    return S_PAYTYPE


async def manual_sale_paytype(update: Update, context: ContextTypes.DEFAULT_TYPE):
    val = update.message.text.strip().lower()
    if "нал" in val and "без" not in val:
        pay = "Наличный"
    elif "без" in val:
        pay = "Безналичный"
    else:
        await update.message.reply_text("⚠️ Выбери: Безналичный или Наличный")
        return S_PAYTYPE
    context.user_data["sale"]["тип_расчета"] = pay
    if pay == "Безналичный":
        kb = ReplyKeyboardMarkup([["Евразийский", "БЦК"], ["⬅️ Назад", "❌ Отмена"]],
                                 resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text("🏦 На какой банк придёт оплата?", reply_markup=kb)
        return S_BANK
    await update.message.reply_text(
        "📦 Фракция? (например: 2-4)\n_Доступны: 0-1, 1-2, 2-4, 4-6, 6-8_",
        parse_mode="Markdown"
    )
    return S_FRAC


async def manual_sale_bank(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Банк поступления при безнал-продаже (для финотчёта)."""
    val = update.message.text.strip().lower()
    if "назад" in val:
        kb = ReplyKeyboardMarkup([["Безналичный", "Наличный"], ["❌ Отмена"]], resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text("💳 Тип расчёта? (Безналичный / Наличный)", reply_markup=kb)
        return S_PAYTYPE
    if "евраз" in val:
        bank = "Евразийский"
    elif "бцк" in val or "bcc" in val:
        bank = "БЦК"
    else:
        await update.message.reply_text("⚠️ Выбери: Евразийский или БЦК")
        return S_BANK
    context.user_data["sale"]["банк"] = bank
    await update.message.reply_text(
        "📦 Фракция? (например: 2-4)\n_Доступны: 0-1, 1-2, 2-4, 4-6, 6-8_",
        parse_mode="Markdown"
    )
    return S_FRAC


async def manual_sale_frac(update: Update, context: ContextTypes.DEFAULT_TYPE):
    frac = update.message.text.strip()
    if not any(f in frac for f in FRACTIONS):
        await update.message.reply_text(
            "⚠️ Не распознал фракцию. Введи одну из: 0-1, 1-2, 2-4, 4-6, 6-8"
        )
        return S_FRAC
    context.user_data["sale"]["фракция"] = frac
    await update.message.reply_text("⚖️ Количество, кг?")
    return S_KG


async def manual_sale_kg(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["sale"]["количество_кг"] = parse_num(update.message.text)
    await update.message.reply_text("💵 Цена за кг, тнг? (как в накладной)")
    return S_PRICE


async def manual_sale_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["sale"]["цена_за_кг"] = parse_num(update.message.text)
    await update.message.reply_text("💰 Сумма с НДС, тнг?")
    return S_SUM_VAT


async def manual_sale_sum_vat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["sale"]["сумма_с_ндс"] = parse_num(update.message.text)
    await update.message.reply_text("💰 Сумма НДС, тнг? (или «-» если нет)")
    return S_VAT


async def manual_sale_vat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["sale"]["сумма_ндс"] = parse_num(update.message.text)
    await update.message.reply_text("📝 Примечание? (или «-» если нет)")
    return S_NOTE


async def manual_sale_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    note = update.message.text.strip()
    context.user_data["sale"]["примечание"] = "" if note in ("-", "—") else note
    d = context.user_data["sale"]
    save_sale(d)
    schedule_refresh()

    report_line = ""
    try:
        ok, msg = write_sale_to_report(d)
        report_line = ("\n\n🧾 Отчёт: " + msg) if ok else ("\n\n⚠️ Отчёт: " + msg)
    except Exception as e:
        logger.error(f"Report write error: {e}")
        report_line = "\n\n⚠️ Отчёт: не удалось записать (склад сохранён)."

    qty_kg = parse_num(d.get("количество_кг", 0))
    reply = (
        f"✅ *Реализация сохранена!*\n\n"
        f"📅 Дата: {d.get('дата', '—')}\n"
        f"🏢 Покупатель: {d.get('покупатель', '—')}\n"
        f"💳 Тип расчёта: {d.get('тип_расчета', '—')}\n"
        f"📦 Фракция: {d.get('фракция', '—')}\n"
        f"⚖️ Количество: {qty_kg} кг\n"
        f"💵 Цена за кг: {d.get('цена_за_кг', 0)} тнг\n"
        f"💰 Сумма с НДС: {d.get('сумма_с_ндс', 0)} тнг\n"
        f"💰 Сумма НДС: {d.get('сумма_ндс', 0)} тнг" + report_line
    )
    await update.message.reply_text(reply, parse_mode="Markdown")
    context.user_data.pop("sale", None)
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.pop("prod", None)
    context.user_data.pop("sale", None)
    context.user_data.pop("pending_prod", None)
    context.user_data.pop("pending_sale", None)
    context.user_data.pop("pending_photo", None)
    context.user_data.pop("exp", None)
    await update.message.reply_text("❌ Отменено.")
    return ConversationHandler.END


# «Спасательный круг»: кнопки главного меню (и «❌ Отмена») срабатывают ВСЕГДА,
# даже посреди незаконченного ввода — сбрасывают его, чтобы бот не залипал.
MENU_ESCAPE_RE = r"^(📸 Производство|📄 Реализация|✍️ Ввод производства|✍️ Ввод реализации|💸 Ввод расхода|❌ Отмена)$"


async def conv_menu_escape(update: Update, context: ContextTypes.DEFAULT_TYPE):
    for k in ("prod", "sale", "exp", "pending_prod", "pending_sale", "pending_photo"):
        context.user_data.pop(k, None)
    if "Отмена" in update.message.text:
        await update.message.reply_text("❌ Отменено, ничего не сохранено.")
    else:
        await update.message.reply_text(
            "⏹ Прошлый незаконченный ввод отменён. Нажми нужную кнопку ещё раз.")
    return ConversationHandler.END


# ---------- Фото и прочий текст ----------

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("⛔ Нет доступа. Обратитесь к администратору.")
        return
    text = update.message.text
    if text == "📸 Производство":
        context.user_data["photo_type"] = "production"
        await update.message.reply_text("📸 Отправь фото отчёта по производству крошки.")
    elif text == "📄 Реализация":
        context.user_data["photo_type"] = "sale"
        await update.message.reply_text("📄 Отправь фото накладной на реализацию.")
    else:
        await update.message.reply_text("Используй кнопки меню или команды /ostatok, /last")


async def _save_sale_from_photo(update, data):
    """Сохранение реализации с фото + попытка записи в отчёт (вынесено для чистоты)."""
    data.setdefault("тип_расчета", "")
    save_sale(data)
    schedule_refresh()

    report_line = ""
    try:
        ok, msg = write_sale_to_report(data)
        report_line = ("\n🧾 Отчёт: " + msg) if ok else ("\n⚠️ Отчёт: " + msg)
    except Exception as e:
        logger.error(f"Report write error: {e}")
        report_line = "\n⚠️ Отчёт: не удалось записать (склад сохранён)."

    qty_kg = parse_num(data.get("количество_кг", 0))
    if not qty_kg and data.get("количество_т"):
        qty_kg = parse_num(data.get("количество_т", 0)) * 1000
    reply = (
        f"✅ *Реализация сохранена!*\n\n"
        f"📅 Дата: {data.get('дата', '—')}\n"
        f"🏢 Покупатель: {data.get('покупатель', '—')}\n"
        f"💳 Тип расчёта: {data.get('тип_расчета') or '— (уточни вручную)'}\n"
        f"📦 Фракция: {data.get('фракция', '—')}\n"
        f"⚖️ Количество: {qty_kg:.0f} кг\n"
        f"💵 Цена за кг: {data.get('цена_за_кг', 0)} тнг\n"
        f"💰 Сумма с НДС: {data.get('сумма_с_ндс', 0)} тнг\n"
        f"💰 Сумма НДС: {data.get('сумма_ндс', 0)} тнг" + report_line
    )
    await update.message.reply_text(reply, parse_mode="Markdown")


async def photo_sale_paytype(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Доспрос нал/безнал после фото-реализации, затем запись в склад и отчёт."""
    val = update.message.text.strip().lower()
    if "нал" in val and "без" not in val:
        pay = "Наличный"
    elif "без" in val:
        pay = "Безналичный"
    else:
        await update.message.reply_text("⚠️ Выбери: Безналичный или Наличный")
        return PHOTO_SALE_PAY
    data = context.user_data.get("pending_sale", {})
    data["тип_расчета"] = pay
    if pay == "Безналичный":
        kb = ReplyKeyboardMarkup([["Евразийский", "БЦК"], ["⬅️ Назад", "❌ Отмена"]],
                                 resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text("🏦 На какой банк придёт оплата?", reply_markup=kb)
        return PHOTO_SALE_BANK
    return await _photo_sale_followup(update, context)


async def _photo_sale_followup(update, context):
    """После типа расчёта/банка доспрашивает недостающее (кол-во, цену, НДС) и сохраняет.
    Нужно для платёжек: там видна только сумма, без кг и цены."""
    d = context.user_data.get("pending_sale", {})
    qty = parse_num(d.get("количество_кг", 0))
    if not qty and d.get("количество_т"):
        qty = parse_num(d.get("количество_т", 0)) * 1000
        d["количество_кг"] = qty
    if not qty:
        kb = ReplyKeyboardMarkup([["❌ Отмена"]], resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text("⚖️ На документе не видно количество. Сколько КГ продали?",
                                        reply_markup=kb)
        return PHOTO_SALE_KG
    total = parse_num(d.get("сумма_с_ндс", 0))
    price = parse_num(d.get("цена_за_кг", 0))
    if not price and not total:
        kb = ReplyKeyboardMarkup([["❌ Отмена"]], resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text("💵 Цена за кг, тнг?", reply_markup=kb)
        return PHOTO_SALE_PRICE
    if not price:
        d["цена_за_кг"] = round(total / qty, 2)
    if not total:
        d["сумма_с_ндс"] = round(qty * parse_num(d.get("цена_за_кг", 0)), 2)
    if not parse_num(d.get("сумма_ндс", 0)) and not d.get("_ндс_решён"):
        total = parse_num(d.get("сумма_с_ндс", 0))
        hint = round(total * 16 / 116, 2)  # НДС в РК с 2026 г. — 16%
        kb = ReplyKeyboardMarkup([[f"В т.ч. НДС 16% = {hint}"], ["Без НДС"], ["❌ Отмена"]],
                                 resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text("💰 Сумма НДС не видна. Выбери кнопкой или введи число:",
                                        reply_markup=kb)
        return PHOTO_SALE_VAT
    d.pop("_ндс_решён", None)
    await _save_sale_from_photo(update, d)
    context.user_data.pop("pending_sale", None)
    context.user_data["photo_type"] = "production"
    return ConversationHandler.END


async def photo_sale_kg(update: Update, context: ContextTypes.DEFAULT_TYPE):
    qty = parse_num(update.message.text)
    if not qty:
        await update.message.reply_text("⚠️ Введи количество числом, в кг.")
        return PHOTO_SALE_KG
    context.user_data.get("pending_sale", {})["количество_кг"] = qty
    return await _photo_sale_followup(update, context)


async def photo_sale_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    price = parse_num(update.message.text)
    if not price:
        await update.message.reply_text("⚠️ Введи цену числом, тнг за кг.")
        return PHOTO_SALE_PRICE
    context.user_data.get("pending_sale", {})["цена_за_кг"] = price
    return await _photo_sale_followup(update, context)


async def photo_sale_vat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    t = update.message.text.strip().lower()
    d = context.user_data.get("pending_sale", {})
    if "без" in t:
        d["сумма_ндс"] = 0
    elif "16%" in t or "т.ч" in t:
        d["сумма_ндс"] = round(parse_num(d.get("сумма_с_ндс", 0)) * 16 / 116, 2)
    else:
        d["сумма_ндс"] = parse_num(update.message.text)
    d["_ндс_решён"] = True
    return await _photo_sale_followup(update, context)


# Виды налогов на ФОТ — порядок важен (как в колонках ведомости и строках финотчёта)
FOT_TAX_KEYS = ["ИПН", "Социальный налог", "Социальные отчисления", "ОПВ", "ОПВР", "ВОСМС", "ОСМС"]


def _norm_period(s):
    """'за Июнь 2026 г.' -> 'июнь 2026' (для дедупликации по примечанию журнала)."""
    low = str(s).lower()
    mon = next((m for m in RU_MONTHS_FULL.values() if m in low), "")
    m = re.search(r"20\d\d", low)
    year = m.group(0) if m else str(datetime.now().year)
    return f"{mon} {year}" if mon else ""


def _payroll_already_entered(period):
    """Сколько уже внесено в журнал по ведомости этого периода, по категориям.
    Нужно для накопительной таблицы: добавляем только дельту, дубли исключены."""
    have = {}
    if not period:
        return have
    try:
        rows = get_expenses_ws().get_all_values()[1:]
    except Exception as e:
        logger.error(f"payroll dedup read error: {e}")
        return have
    for row in rows:
        if len(row) < 8:
            row = row + [""] * (8 - len(row))
        note = str(row[7]).lower()
        if "ведомост" not in note or period not in note:
            continue
        cat = str(row[3]).strip()
        have[cat] = have.get(cat, 0) + parse_num(row[1])
    return have


async def recognize_payroll(image_bytes: bytes):
    """Распознаёт расчётную ведомость: зарплата по разделам + оплаченные налоги ФОТ."""
    image_b64 = base64.b64encode(image_bytes).decode()
    prompt = (
        "Это фото расчётной ведомости организации по заработной плате (возможно, с колонками налогов).\n"
        "Верни ТОЛЬКО JSON без markdown:\n"
        '{"месяц":"месяц и год из заголовка","ауп":0,"производство":0,"итого":0,'
        '"банк":"банк из шапки колонки «Перечислено в … банк», если назван",'
        '"налоги":{"ИПН":0,"Социальный налог":0,"Социальные отчисления":0,"ОПВ":0,"ОПВР":0,"ВОСМС":0,"ОСМС":0}}\n'
        "ауп — ИТОГ раздела «Основное подразделение» (АУП) по колонке «Перечислено в банк»; "
        "производство — ИТОГ раздела «Первая линия»/ПРОИЗВОДСТВО той же колонки; итого — общий итог. "
        "налоги — из строки ИТОГО (или строки организации) по колонкам ИПН, Социальный налог, "
        "Социальные отчисления, ОПВ, ОПВР, ВОСМС, ООСМС (=ОСМС). Пустая колонка или её нет — 0. "
        "Бери именно итоговые строки, не отдельных сотрудников."
    )
    return await call_claude(image_b64, prompt)


async def payroll_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    t = update.message.text.strip().lower()
    d = context.user_data.get("pending_payroll", {})
    if t in ("сегодня", "today"):
        d["дата"] = datetime.now().strftime("%d.%m.%Y")
    else:
        dt = parse_date_or_none(t)
        if dt is None:
            await update.message.reply_text("⚠️ Не понял дату. Введи дд.мм.гггг или «сегодня».")
            return PAYROLL_DATE
        d["дата"] = date_to_str(dt)
    if not d.get("банк"):
        kb = ReplyKeyboardMarkup([["Евразийский", "БЦК"], ["❌ Отмена"]],
                                 resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text("🏦 С какого банка перечислена зарплата?", reply_markup=kb)
        return PAYROLL_BANK
    return await _payroll_confirm_ask(update, context)


async def payroll_bank(update: Update, context: ContextTypes.DEFAULT_TYPE):
    val = update.message.text.strip().lower()
    if "евраз" in val:
        bank = "Евразийский"
    elif "бцк" in val or "bcc" in val:
        bank = "БЦК"
    else:
        await update.message.reply_text("⚠️ Выбери: Евразийский или БЦК")
        return PAYROLL_BANK
    context.user_data.get("pending_payroll", {})["банк"] = bank
    return await _payroll_confirm_ask(update, context)


async def _payroll_confirm_ask(update, context):
    d = context.user_data.get("pending_payroll", {})
    items = "\n".join(f"• {i['категория']}: {round(parse_num(i['сумма']))} тнг"
                      for i in d.get("add", []))
    t_add = sum(parse_num(i["сумма"]) for i in d.get("add", []))
    kb = ReplyKeyboardMarkup([["✅ Да, сохранить"], ["❌ Отмена"]],
                             resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text(
        f"Проверь перед сохранением (ведомость за {d.get('месяц') or '—'}):\n\n"
        f"📅 Дата оплаты: {d.get('дата', '—')}\n"
        f"🏦 Банк: {d.get('банк', '—')}\n\n"
        f"Добавляю в журнал:\n{items}\n\n"
        f"💰 Итого добавляется: {round(t_add)} тнг. Всё разнесётся по отчётам само.",
        reply_markup=kb)
    return PAYROLL_CONFIRM


async def payroll_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ans = update.message.text.strip().lower()
    d = context.user_data.get("pending_payroll", {})
    if "да" in ans or "сохран" in ans:
        base = {"дата": d.get("дата", ""), "источник": d.get("банк", ""), "контрагент": "",
                "примечание": f"по ведомости за {d.get('месяц') or '—'}"}
        for item in d.get("add", []):
            save_expense({**base, **item})
        schedule_refresh()
        n = len(d.get("add", []))
        context.user_data.pop("pending_payroll", None)
        await update.message.reply_text(
            f"✅ Сохранено строк: {n}. Финотчёт, «Калькуляция» и сводные обновятся сами. "
            f"Пришлёшь эту же ведомость позже с новыми оплаченными налогами — добавлю только новое.")
        return ConversationHandler.END
    context.user_data.pop("pending_payroll", None)
    await update.message.reply_text("❌ Отменено, ничего не сохранено.")
    return ConversationHandler.END


# ---------- Табель учёта рабочего времени ----------
SHEET_TABEL = "Табель"
TABEL_HEADERS = ["Дата", "Сотрудник", "Должность", "Часы", "Примечание"]


def get_tabel_ws():
    return get_worksheet(SHEET_TABEL, TABEL_HEADERS)


def _tabel_existing(month, year):
    """Ключи (день, сотрудник) уже внесённых записей табеля за месяц — для дедупликации."""
    have = set()
    try:
        rows = get_tabel_ws().get_all_values()[1:]
    except Exception as e:
        logger.error(f"tabel read error: {e}")
        return have
    for row in rows:
        if len(row) < 2:
            continue
        dt = parse_date_or_none(row[0])
        if not dt or dt.month != month or dt.year != year:
            continue
        have.add((dt.day, _fin_norm(row[1])))
    return have


def _tabel_period_nums(period):
    """'июнь 2026' -> (6, 2026) или None."""
    if not period:
        return None
    parts = period.split()
    mon = next((n for n, nm in RU_MONTHS_FULL.items() if nm == parts[0]), None)
    if not mon:
        return None
    year = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else datetime.now().year
    return mon, year


async def recognize_tabel(image_bytes: bytes):
    """Распознаёт табель: сотрудники × дни месяца, отработанные часы."""
    image_b64 = base64.b64encode(image_bytes).decode()
    prompt = (
        "Это фото табеля учёта рабочего времени: строки — сотрудники, колонки — дни месяца "
        "(1–31), в ячейках отработанные часы (число) или отметки (В=выходной, О=отпуск, "
        "Б=больничный — такие дни пропускай).\n"
        "Верни ТОЛЬКО JSON без markdown:\n"
        '{"месяц":"месяц и год из заголовка","сотрудники":[{"фио":"...","должность":"...",'
        '"дни":{"1":8,"2":8}}]}\n'
        "В «дни» включай ТОЛЬКО рабочие дни (ключ — номер дня, значение — часы). Если стоит "
        "отметка выхода (Я, 1, +) без числа часов — ставь 8. Что не читается — пропусти."
    )
    return await call_claude(image_b64, prompt)


async def _tabel_prepare(update, context):
    """Считает НОВЫЕ записи табеля (дедуп по дню+сотруднику) и просит подтверждение."""
    rawlist = context.user_data.get("pending_tabel_raw") or []
    period = context.user_data.get("pending_tabel_period", "")
    nums = _tabel_period_nums(period)
    if not nums:
        await update.message.reply_text(
            "📅 Не понял, за какой месяц табель. Напиши, например: «июнь 2026».")
        return TABEL_MONTH
    mon, year = nums
    have = await asyncio.to_thread(_tabel_existing, mon, year)
    rows, skipped, hours, emps = [], 0, 0.0, set()
    for emp in rawlist:
        fio = str(emp.get("фио", "")).strip()
        if not fio:
            continue
        role = str(emp.get("должность", "")).strip()
        days = emp.get("дни") if isinstance(emp.get("дни"), dict) else {}
        for dstr, h in days.items():
            try:
                day = int(str(dstr).strip())
            except Exception:
                continue
            hv = parse_num(h)
            if day < 1 or day > 31 or hv <= 0:
                continue
            if (day, _fin_norm(fio)) in have:
                skipped += 1
                continue
            emps.add(fio)
            hours += hv
            rows.append([f"{day:02d}.{mon:02d}.{year}", fio, role, hv, ""])
    if not rows:
        if skipped:
            msg = f"📅 Табель за {period}: новых записей нет, всё уже внесено (пропущено {skipped})."
        else:
            msg = "📅 В табеле не нашёл ни одного дня с часами. Переснимай чётче."
        await update.message.reply_text(msg)
        context.user_data.pop("pending_tabel_raw", None)
        context.user_data.pop("pending_tabel_period", None)
        return ConversationHandler.END
    rows.sort(key=lambda r: (parse_date_or_none(r[0]) or datetime.min, r[1]))
    context.user_data["pending_tabel_rows"] = rows
    note = f"\n♻️ Уже было внесено (пропускаю): {skipped}" if skipped else ""
    kb = ReplyKeyboardMarkup([["✅ Да, сохранить"], ["❌ Отмена"]],
                             resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text(
        f"📅 Табель за {period}:\n"
        f"👥 Сотрудников: {len(emps)}\n"
        f"🆕 Новых записей (человеко-дней): {len(rows)}\n"
        f"⏱ Часов добавится: {round(hours)}{note}\n\nЗаписать в лист «Табель»?",
        reply_markup=kb)
    return TABEL_CONFIRM


async def tabel_month(update: Update, context: ContextTypes.DEFAULT_TYPE):
    period = _norm_period(update.message.text)
    if not period:
        await update.message.reply_text("⚠️ Напиши месяц словом, например: «июнь 2026».")
        return TABEL_MONTH
    context.user_data["pending_tabel_period"] = period
    return await _tabel_prepare(update, context)


async def tabel_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ans = update.message.text.strip().lower()
    rows = context.user_data.get("pending_tabel_rows") or []
    for k in ("pending_tabel_rows", "pending_tabel_raw", "pending_tabel_period"):
        context.user_data.pop(k, None)
    if ("да" in ans or "сохран" in ans) and rows:
        ws = await asyncio.to_thread(get_tabel_ws)
        await asyncio.to_thread(ws.append_rows, rows, value_input_option="USER_ENTERED")
        await update.message.reply_text(
            f"✅ Табель записан: строк {len(rows)}. Смотреть — в кабинете, раздел «Табель».")
        return ConversationHandler.END
    await update.message.reply_text("❌ Отменено, ничего не сохранено.")
    return ConversationHandler.END


async def photo_sale_bank(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Доспрос банка после фото-реализации (безнал) — для финотчёта."""
    val = update.message.text.strip().lower()
    if "назад" in val:
        kb = ReplyKeyboardMarkup([["Безналичный", "Наличный"], ["❌ Отмена"]], resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text("💳 Тип расчёта? (Безналичный / Наличный)", reply_markup=kb)
        return PHOTO_SALE_PAY
    if "евраз" in val:
        bank = "Евразийский"
    elif "бцк" in val or "bcc" in val:
        bank = "БЦК"
    else:
        await update.message.reply_text("⚠️ Выбери: Евразийский или БЦК")
        return PHOTO_SALE_BANK
    data = context.user_data.get("pending_sale", {})
    data["банк"] = bank
    return await _photo_sale_followup(update, context)


async def detect_doc_type(image_bytes: bytes):
    """Определяет по фото: накладная-реализация, отчёт производства или расходный документ."""
    image_b64 = base64.b64encode(image_bytes).decode()
    prompt = (
        "Определи тип документа на фото. Это одно из трёх:\n"
        "- НАКЛАДНАЯ на отпуск/реализацию: компания ПРОДАЁТ резиновую крошку — есть покупатель, "
        "фракции крошки, цены, суммы, НДС;\n"
        "- ОТЧЁТ ПРОИЗВОДСТВА крошки: смена, оператор, вес шин, фракции в кг, мешки;\n"
        "- РАСХОДНЫЙ ДОКУМЕНТ: платёжное поручение, счёт на оплату, чек, квитанция — компания "
        "ПЛАТИТ поставщику. Если плательщик — ТОО «Еділ и компания», это расход;\n"
        "- РАСЧЁТНАЯ ВЕДОМОСТЬ по заработной плате: список сотрудников с должностями, разделы "
        "АУП и ПРОИЗВОДСТВО, колонка «перечислено в банк», итоги;\n"
        "- ТАБЕЛЬ учёта рабочего времени: строки — сотрудники, колонки — дни месяца (1..31), "
        "в ячейках часы или отметки выходов.\n"
        'Верни ТОЛЬКО JSON: {"тип":"реализация"} или {"тип":"производство"} или {"тип":"расход"} '
        'или {"тип":"ведомость"} или {"тип":"табель"}.'
    )
    try:
        res = await call_claude(image_b64, prompt)
        t = str(res.get("тип", "")).strip().lower() if isinstance(res, dict) else ""
    except Exception as e:
        logger.error(f"detect_doc_type error: {e}")
        t = ""
    if "табел" in t:
        return "tabel"
    if "ведомост" in t:
        return "payroll"
    if "расход" in t:
        return "expense"
    return "sale" if "реализ" in t else "production"


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Приём фото: определяем тип (реализация/производство) и просим подтвердить."""
    if not is_allowed(update):
        await update.message.reply_text("⛔ Нет доступа. Обратитесь к администратору.")
        return ConversationHandler.END
    await update.message.reply_text("📸 Получил фото, определяю тип документа...")
    try:
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        image_bytes = bytes(await file.download_as_bytearray())
        doc_type = await detect_doc_type(image_bytes)
        context.user_data["pending_photo"] = {"file_id": photo.file_id, "type": doc_type}
        labels = {"sale": "📄 РЕАЛИЗАЦИЯ (накладная)",
                  "production": "📸 ПРОИЗВОДСТВО (отчёт)",
                  "expense": "💸 РАСХОД (платёжка / счёт / чек)",
                  "payroll": "🧾 ВЕДОМОСТЬ ФОТ (зарплата)",
                  "tabel": "📅 ТАБЕЛЬ (рабочее время)"}
        alt = {"sale": "🔁 Это реализация", "production": "🔁 Это производство",
               "expense": "🔁 Это расход", "payroll": "🔁 Это ведомость ФОТ",
               "tabel": "🔁 Это табель"}
        others = [alt[k] for k in ("sale", "production", "expense", "payroll", "tabel")
                  if k != doc_type]
        label = labels[doc_type]
        kb = ReplyKeyboardMarkup([["✅ Да, верно"], others[:2], others[2:], ["❌ Отмена"]],
                                 resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text(
            f"Определил: {label}\n\nВерно? Подтверди — распознаю и сохраню.",
            reply_markup=kb,
        )
        return PHOTO_TYPE_CONFIRM
    except Exception as e:
        logger.error(f"handle_photo detect error: {e}")
        await update.message.reply_text("❌ Не смог обработать фото. Попробуй ещё раз.")
        return ConversationHandler.END


async def photo_type_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Подтверждение типа: Да — сохраняем; Нет — берём другой тип."""
    ans = update.message.text.strip().lower()
    pending = context.user_data.get("pending_photo")
    if not pending:
        await update.message.reply_text("Пришли фото заново, пожалуйста.")
        return ConversationHandler.END
    if "да" in ans or "верно" in ans:
        photo_type = pending["type"]
    elif "табел" in ans:
        photo_type = "tabel"
    elif "ведомост" in ans:
        photo_type = "payroll"
    elif "расход" in ans:
        photo_type = "expense"
    elif "производ" in ans:
        photo_type = "production"
    elif "реализ" in ans:
        photo_type = "sale"
    else:
        await update.message.reply_text("Нажми «✅ Да, верно», «🔁 Это …» или «❌ Отмена».")
        return PHOTO_TYPE_CONFIRM
    context.user_data.pop("pending_photo", None)
    await update.message.reply_text("🔎 Принято, распознаю данные...")
    try:
        file = await context.bot.get_file(pending["file_id"])
        image_bytes = bytes(await file.download_as_bytearray())
    except Exception as e:
        logger.error(f"photo_type_confirm download error: {e}")
        await update.message.reply_text("❌ Не смог скачать фото. Пришли заново.")
        return ConversationHandler.END
    return await _dispatch_photo(update, context, photo_type, image_bytes)


async def _dispatch_photo(update, context, photo_type, image_bytes):
    """Распознаёт и сохраняет фото по подтверждённому типу."""
    try:
        if photo_type == "tabel":
            await update.message.reply_text("📅 Распознаю табель...")
            data = await recognize_tabel(image_bytes)
            logger.info(f"Tabel data: {data}")
            rawlist = data.get("сотрудники") if isinstance(data, dict) else None
            if not rawlist:
                await update.message.reply_text("❌ Не смог прочитать табель. Переснимай чётче.")
                return ConversationHandler.END
            context.user_data["pending_tabel_raw"] = rawlist
            context.user_data["pending_tabel_period"] = _norm_period(str(data.get("месяц", "")))
            return await _tabel_prepare(update, context)

        if photo_type == "payroll":
            await update.message.reply_text("🧾 Распознаю ведомость ФОТ...")
            data = await recognize_payroll(image_bytes)
            logger.info(f"Payroll data: {data}")
            aup = parse_num(data.get("ауп", 0))
            prod = parse_num(data.get("производство", 0))
            total = parse_num(data.get("итого", 0))
            taxes = data.get("налоги") if isinstance(data.get("налоги"), dict) else {}
            if not aup and not prod and not any(parse_num(v) for v in taxes.values()):
                await update.message.reply_text(
                    "❌ Не смог прочитать итоги с ведомости (разделы и налоги пустые). Переснимай чётче.")
                return ConversationHandler.END
            warn = ""
            if total and round(aup + prod) != round(total):
                warn = (f"\n⚠️ Сверка не сошлась: АУП {round(aup)} + Производство {round(prod)} "
                        f"= {round(aup + prod)}, а «Итого» на бланке {round(total)}. Проверь внимательно!")
            period = _norm_period(str(data.get("месяц", "")))
            bank_raw = str(data.get("банк", "")).lower()
            bank = ("Евразийский" if "евраз" in bank_raw
                    else "БЦК" if ("бцк" in bank_raw or "центркредит" in bank_raw) else "")
            # таблица накопительная: сравниваем с уже внесённым, добавляем только дельту
            have = await asyncio.to_thread(_payroll_already_entered, period)
            targets = [("ФОТ производственный", "Производство", prod),
                       ("ФОТ административный", "Администрация", aup)]
            for vid in FOT_TAX_KEYS:
                targets.append((f"Налог ФОТ: {vid}", "Администрация", parse_num(taxes.get(vid, 0))))
            add, lines = [], []
            for cat, grp, target in targets:
                if target <= 0:
                    continue
                done = have.get(cat, 0)
                delta = round(target - done, 2)
                if delta > 0.5:
                    add.append({"категория": cat, "группа": grp, "сумма": delta})
                    mark = f"➕ добавлю {round(delta)}" + (f" (внесено {round(done)})" if done else "")
                elif delta < -0.5:
                    mark = f"⚠️ в таблице МЕНЬШЕ внесённого ({round(target)} < {round(done)}) — не трогаю"
                else:
                    mark = "✔ уже внесено"
                lines.append(f"• {cat}: {round(target)} — {mark}")
            head = f"🧾 Ведомость за {period or '—'}:\n" + "\n".join(lines) + warn
            if not add:
                await update.message.reply_text(head + "\n\n✅ Нового нет — ничего не добавляю.")
                return ConversationHandler.END
            context.user_data["pending_payroll"] = {"месяц": period, "банк": bank, "add": add}
            await update.message.reply_text(
                head + "\n\n📅 Какой ДАТОЙ оплачено добавляемое? (дд.мм.гггг или «сегодня»)")
            return PAYROLL_DATE

        if photo_type == "expense":
            await update.message.reply_text("💸 Распознаю расходный документ...")
            data = await recognize_expense(image_bytes)
            logger.info(f"Expense photo data: {data}")
            return await _exp_from_photo_data(update, context, data)

        if photo_type == "sale":
            data = await recognize_sale(image_bytes)
            logger.info(f"Sale data: {data}")
            # тип расчёта (нал/безнал) на накладной обычно не виден — нормализуем распознанное
            pay_raw = str(data.get("тип_расчета", "")).strip().lower()
            if "нал" in pay_raw and "без" not in pay_raw:
                data["тип_расчета"] = "Наличный"
            elif "без" in pay_raw:
                data["тип_расчета"] = "Безналичный"
            else:
                data["тип_расчета"] = ""
            if not data["тип_расчета"]:
                # доспрашиваем, чтобы продажа попала и в отчёт «Реализация 2026г»
                context.user_data["pending_sale"] = data
                kb = ReplyKeyboardMarkup([["Безналичный", "Наличный"], ["❌ Отмена"]],
                                         resize_keyboard=True, one_time_keyboard=True)
                await update.message.reply_text(
                    "📄 Накладную распознал. Уточни тип расчёта, чтобы продажа попала и в отчёт:",
                    reply_markup=kb,
                )
                return PHOTO_SALE_PAY
            # тип расчёта распознан с фото: банк (для безнала) + доспрос недостающего
            context.user_data["pending_sale"] = data
            if data["тип_расчета"] == "Безналичный":
                kb = ReplyKeyboardMarkup([["Евразийский", "БЦК"], ["⬅️ Назад", "❌ Отмена"]],
                                         resize_keyboard=True, one_time_keyboard=True)
                await update.message.reply_text("🏦 На какой банк придёт оплата?", reply_markup=kb)
                return PHOTO_SALE_BANK
            return await _photo_sale_followup(update, context)

        # --- производство ---
        data = await recognize_production(image_bytes)
        logger.info(f"Production data: {data}")
        records = data if isinstance(data, list) else [data]

        # Проверяем дату КАЖДОЙ записи. Claude должен был отдать дд.мм.гггг.
        # Если дата не читается — НЕ пишем молча, а просим ввести вручную.
        bad_date = []
        for r in records:
            dt = parse_date_or_none(r.get("дата", ""))
            if dt is None:
                bad_date.append(r)
            else:
                r["дата"] = date_to_str(dt)  # приводим к единому виду дд.мм.гггг

        if bad_date:
            await update.message.reply_text(
                "⚠️ Не смог уверенно прочитать ДАТУ на фото.\n\n"
                "Чтобы не записать неверно, внеси эту смену через «✍️ Ввод производства» "
                "(там дату введёшь вручную), или переснимай фото так, чтобы дата была чёткой "
                "и с точками (например 05.06.2026).\n\n"
                "В склад с этого фото ничего не записал."
            )
            return ConversationHandler.END

        # Сверка: сумма по фракциям должна совпасть с «всего», написанным на бланке
        mismatch = []
        for r in records:
            written = parse_num(r.get("всего_итог", 0))
            calc = sum(parse_num(r.get(fk, 0)) for fk in
                       ("фракция_0_1", "фракция_1_2", "фракция_2_4", "фракция_4_6", "фракция_6_8"))
            if written and abs(calc - written) > 1:
                mismatch.append((calc, written))
        if mismatch:
            calc, written = mismatch[0]
            await update.message.reply_text(
                "⚠️ Расхождение по сумме крошки.\n\n"
                f"На бланке «всего»: {written:.0f} кг\n"
                f"По распознанным фракциям: {calc:.0f} кг\n\n"
                "Возможно, цифра распозналась неверно. Проверь фото и внеси смену вручную "
                "через «✍️ Ввод производства», или переснимай чётче.\n\n"
                "В склад с этого фото ничего не записал."
            )
            return ConversationHandler.END

        # Делим на «чистые» (сразу пишем) и «вероятные дубли» (спросим)
        clean, dups = [], []
        for r in records:
            try:
                if find_duplicate_production(r):
                    dups.append(r)
                else:
                    clean.append(r)
            except Exception as e:
                logger.error(f"dup check error (photo): {e}")
                clean.append(r)  # если проверка упала — не теряем данные, пишем

        # Чистые записи сохраняем сразу
        for item in clean:
            save_production(item)
        if clean:
            schedule_refresh()

        # Если есть подозрение на дубль — спрашиваем по первому, остальные дубли держим в очереди
        if dups:
            context.user_data["pending_prod"] = dups
            first_dup = dups[0]
            sig_total = sum(_fractions_signature(first_dup))
            saved_note = ""
            if clean:
                saved_note = f"\n\n(Заодно сохранил новых записей: {len(clean)}.)"
            kb = ReplyKeyboardMarkup([["Да, добавить", "Нет, пропустить"]],
                                     resize_keyboard=True, one_time_keyboard=True)
            await update.message.reply_text(
                f"⚠️ Похоже, эта запись УЖЕ ЕСТЬ в складе:\n\n"
                f"📅 Дата: {first_dup.get('дата', '—')}\n"
                f"👤 Оператор: {first_dup.get('фио', '—')}\n"
                f"⚖️ Всего крошки: {sig_total:.0f} кг\n\n"
                f"Добавить её всё равно?" + saved_note,
                reply_markup=kb,
                parse_mode="Markdown"
            )
            return PHOTO_PROD_CONFIRM

        # Дублей нет — обычный отчёт
        if not clean:
            await update.message.reply_text("⚠️ С фото не удалось получить данные. Попробуй ещё раз.")
            return ConversationHandler.END

        first = clean[-1]
        await update.message.reply_text(
            _prod_summary(first).replace("сохранено!", f"сохранено! (новых: {len(clean)})"),
            parse_mode="Markdown"
        )
        return ConversationHandler.END

    except Exception as e:
        logger.error(f"Error: {e}")
        await update.message.reply_text(
            "❌ Не смог распознать фото. Попробуй:\n"
            "• Сделать фото чётче\n"
            "• Хорошее освещение\n"
            "• Всё в кадре"
        )
        return ConversationHandler.END


async def photo_prod_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """NEW: ответ Да/Нет на вопрос о дубле фото-производства."""
    ans = update.message.text.strip().lower()
    dups = context.user_data.get("pending_prod", [])
    context.user_data.pop("pending_prod", None)

    if not dups:
        await update.message.reply_text("Нечего подтверждать. Отправь фото заново.")
        return ConversationHandler.END

    if "да" in ans:
        for item in dups:
            save_production(item)
        schedule_refresh()
        await update.message.reply_text(
            f"✅ Добавил, несмотря на совпадение. Записей: {len(dups)}.\n"
            f"Проверь остаток: /ostatok"
        )
    else:
        await update.message.reply_text(
            "👍 Пропустил — дубль в склад не попал."
        )
    return ConversationHandler.END


# ---------- Автобэкап данных в закрытый Telegram-канал ----------

def _build_backup_xlsx():
    """Собирает один .xlsx: листы склада (Производство, Реализация) + вкладки отчёта реализации.
    Возвращает (bytes_файла, counts) — counts это число строк данных по ключевым листам."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    counts = {}
    book = get_spreadsheet()
    for name in ("Производство", "Реализация"):
        try:
            rows = book.worksheet(name).get_all_values()
        except Exception as e:
            logger.error(f"backup read {name}: {e}")
            rows = [["(не удалось прочитать лист)"]]
        sheet = wb.create_sheet(name[:31])
        for row in rows:
            sheet.append(row)
        counts[name] = max(0, len(rows) - 1)
    try:
        rep = get_report_spreadsheet()
        for ws in rep.worksheets():
            rows = ws.get_all_values()
            sheet = wb.create_sheet(("Отчёт-" + ws.title)[:31])
            for row in rows:
                sheet.append(row)
    except Exception as e:
        logger.error(f"backup read report: {e}")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), counts


async def do_backup(bot):
    """Собирает и отправляет бэкап в закрытый канал. Возвращает 'ok' или текст ошибки."""
    if not BACKUP_CHANNEL_ID:
        logger.warning("BACKUP_CHANNEL_ID не задан — бэкап пропущен")
        return "не задан BACKUP_CHANNEL_ID"
    data, counts = await asyncio.to_thread(_build_backup_xlsx)
    global _last_backup_counts
    warn = ""
    for key in ("Производство", "Реализация"):
        old = _last_backup_counts.get(key)
        new = counts.get(key, 0)
        if old is not None and (old - new) > 2:
            warn += f"\n⚠️ В «{key}» строк стало МЕНЬШЕ: было {old}, стало {new}. Проверь, не удалили ли случайно!"
    _last_backup_counts = counts
    now = datetime.now(KZ_TZ)
    fname = "edil_backup_" + now.strftime("%Y-%m-%d_%H%M") + ".xlsx"
    caption = ("🗄 Бэкап Едил · " + now.strftime("%d.%m.%Y %H:%M")
               + f"\nПроизводство: {counts.get('Производство', 0)} · Реализация: {counts.get('Реализация', 0)} строк")
    if warn:
        caption += "\n" + warn
    await bot.send_document(
        chat_id=int(BACKUP_CHANNEL_ID),
        document=InputFile(io.BytesIO(data), filename=fname),
        caption=caption,
    )
    return "ok"


async def _backup_loop(application):
    """Раз в сутки в BACKUP_HOUR по местному времени (UTC+5) делает бэкап."""
    await asyncio.sleep(10)
    while True:
        now = datetime.now(KZ_TZ)
        target = now.replace(hour=BACKUP_HOUR, minute=0, second=0, microsecond=0)
        if target <= now:
            target += timedelta(days=1)
        await asyncio.sleep(max(1, (target - now).total_seconds()))
        try:
            await do_backup(application.bot)
        except Exception as e:
            logger.error(f"scheduled backup error: {e}")
        try:
            await asyncio.to_thread(_build_svod)
            await asyncio.to_thread(_build_sebes)
        except Exception as e:
            logger.error(f"scheduled svod error: {e}")
        try:
            await _run_refresh()  # страховочный суточный прогон «Калькуляции» и финотчёта
        except Exception as e:
            logger.error(f"scheduled refresh error: {e}")


async def _post_init(application):
    application.create_task(_backup_loop(application))


async def cmd_backup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ручной запуск бэкапа для проверки. Только для разрешённых пользователей."""
    if not is_allowed(update):
        await update.message.reply_text("⛔ Нет доступа.")
        return
    await update.message.reply_text("🗄 Делаю бэкап, пара секунд...")
    try:
        res = await do_backup(context.bot)
    except Exception as e:
        logger.error(f"manual backup error: {e}")
        await update.message.reply_text(f"❌ Ошибка бэкапа: {e}")
        return
    if res == "ok":
        await update.message.reply_text("✅ Бэкап отправлен в канал.")
    else:
        await update.message.reply_text(f"⚠️ {res}")


# ---------- Журнал «Расходы»: ввод расходов компании ----------

SHEET_EXPENSES = "Расходы"
EXPENSE_HEADERS = ["Дата", "Сумма", "Группа", "Категория", "Нал/Безнал", "Источник", "Контрагент", "Примечание"]
EXPENSE_GROUPS = {
    "Производство": [
        "Материалы и запчасти", "Упаковка (мешки)", "Электроэнергия", "ГСМ",
        "ФОТ производственный", "Транспорт и логистика", "Лизинг (вознаграждение)",
        "Амортизация производства",
    ],
    "Администрация": [
        "ФОТ административный", "Налоги", "Услуги банка", "Страхование", "Аудит",
        "Прочие административные", "Амортизация административная",
    ],
    "Капзатраты": [
        "Основные средства", "Лизинг (основной долг)", "Прочие капзатраты",
    ],
}
EXPENSE_SOURCES = ["Евразийский", "БЦК", "Касса 1", "Касса 2", "Неденежный"]


def get_expenses_ws():
    return get_worksheet(SHEET_EXPENSES, EXPENSE_HEADERS)


def _source_to_paytype(src):
    if src in ("Евразийский", "БЦК"):
        return "Безналичный"
    if src in ("Касса 1", "Касса 2"):
        return "Наличный"
    return "Неденежный"


def save_expense(data):
    ws = get_expenses_ws()
    src = data.get("источник", "")
    row = [
        data.get("дата", datetime.now().strftime("%d.%m.%Y")),
        parse_num(data.get("сумма", 0)),
        data.get("группа", ""),
        data.get("категория", ""),
        _source_to_paytype(src),
        src,
        data.get("контрагент", ""),
        data.get("примечание", ""),
    ]
    ws.append_row(row, value_input_option="USER_ENTERED")


def _exp_summary(d, saved=True):
    head = "✅ *Расход сохранён!*" if saved else "Проверь расход перед сохранением:"
    return (
        f"{head}\n\n"
        f"📅 Дата: {d.get('дата', '—')}\n"
        f"💵 Сумма: {d.get('сумма', 0)} тнг\n"
        f"📂 Группа: {d.get('группа', '—')}\n"
        f"🏷 Категория: {d.get('категория', '—')}\n"
        f"🏦 Источник: {d.get('источник', '—')}\n"
        f"🏢 Контрагент: {d.get('контрагент') or '—'}\n"
        f"📝 Примечание: {d.get('примечание') or '—'}"
    )


# --- Справочник поставщиков (база) ---
SHEET_SUPPLIERS = "Поставщики"
SUPPLIER_HEADERS = ["Поставщик"]
SUPPLIER_FUZZY = 0.88
_LEGAL_PREFIXES = ("товарищество с ограниченной ответственностью", "индивидуальный предприниматель",
                   "тоо", "ип", "ао", "оао", "зао", "ооо", "чп", "кх")


def _normalize_supplier(name):
    s = str(name).strip().lower()
    for ch in "«»\"'":
        s = s.replace(ch, " ")
    for p in _LEGAL_PREFIXES:
        s = re.sub(r"\b" + re.escape(p) + r"\b", " ", s)
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def get_suppliers_ws():
    return get_worksheet(SHEET_SUPPLIERS, SUPPLIER_HEADERS)


def load_suppliers():
    try:
        vals = get_suppliers_ws().col_values(1)[1:]
    except Exception as e:
        logger.error(f"load_suppliers: {e}")
        return []
    return [str(v).strip() for v in vals if str(v).strip()]


def add_supplier(name):
    name = str(name).strip()
    if not name:
        return
    try:
        get_suppliers_ws().append_row([name], value_input_option="USER_ENTERED")
    except Exception as e:
        logger.error(f"add_supplier: {e}")


def match_supplier(name):
    """Возвращает (status, canonical): 'exact' | 'fuzzy' | 'new'."""
    import difflib
    norm = _normalize_supplier(name)
    if not norm:
        return ("new", name)
    suppliers = load_suppliers()
    for s in suppliers:
        if _normalize_supplier(s) == norm:
            return ("exact", s)
    best, best_r = None, 0.0
    for s in suppliers:
        r = difflib.SequenceMatcher(None, norm, _normalize_supplier(s)).ratio()
        if r > best_r:
            best, best_r = s, r
    if best is not None and best_r >= SUPPLIER_FUZZY:
        return ("fuzzy", best)
    return ("new", name)


async def _supplier_resolve(update, context, name):
    """Сверка поставщика. Возвращает следующий state или None (можно идти дальше)."""
    context.user_data["exp"]["контрагент"] = name
    if not name:
        return None
    st, cand = match_supplier(name)
    if st == "exact":
        context.user_data["exp"]["контрагент"] = cand
        return None
    if st == "fuzzy":
        context.user_data["exp"]["_sup_cand"] = cand
        kb = ReplyKeyboardMarkup([["✅ Да, это он", "🆕 Нет, другой"]],
                                 resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text(f"🏢 Похоже на «{cand}» из базы. Это он?", reply_markup=kb)
        return E_SUP_CONFIRM
    kb = ReplyKeyboardMarkup([["✅ Да, записать", "✏️ Исправить имя"]],
                             resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text(
        f"🆕 Новый поставщик: «{name}». Проверь, правильно ли имя. Записать в базу?", reply_markup=kb)
    return E_SUP_NEW


async def _supplier_done(update, context):
    """Дальше после поставщика: фото → группа, ручной → примечание."""
    if context.user_data.get("exp", {}).get("_photo"):
        kb = ReplyKeyboardMarkup([["Производство"], ["Администрация"], ["Капзатраты"], ["❌ Отмена"]],
                                 resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text("Выбери группу расхода:", reply_markup=kb)
        return E_GROUP
    await update.message.reply_text("📝 Примечание? Или «-», если нет.")
    return E_NOTE


async def exp_sup_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ответ на «похоже на X»."""
    ans = update.message.text.strip().lower()
    exp = context.user_data.get("exp", {})
    if "да" in ans:
        exp["контрагент"] = exp.get("_sup_cand", exp.get("контрагент", ""))
        exp.pop("_sup_cand", None)
        return await _supplier_done(update, context)
    exp.pop("_sup_cand", None)
    name = exp.get("контрагент", "")
    kb = ReplyKeyboardMarkup([["✅ Да, записать", "✏️ Исправить имя"]],
                             resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text(
        f"🆕 Новый поставщик: «{name}». Проверь, правильно ли имя. Записать в базу?", reply_markup=kb)
    return E_SUP_NEW


async def exp_sup_new(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Подтверждение нового поставщика перед записью в базу."""
    ans = update.message.text.strip().lower()
    exp = context.user_data.get("exp", {})
    if "да" in ans or "запис" in ans:
        add_supplier(exp.get("контрагент", ""))
        return await _supplier_done(update, context)
    await update.message.reply_text("✏️ Введи правильное имя поставщика:")
    return E_SUP_EDIT


async def exp_sup_edit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Исправленное имя → повторная сверка."""
    name = update.message.text.strip()
    nxt = await _supplier_resolve(update, context, name)
    if nxt is not None:
        return nxt
    return await _supplier_done(update, context)


async def manual_exp_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("⛔ Нет доступа.")
        return ConversationHandler.END
    context.user_data["exp"] = {}
    await update.message.reply_text("💸 Внесём расход. Пришли ФОТО счёта/чека — распознаю сам, или введи дату вручную (дд.мм.гггг или «сегодня»).")
    return E_DATE


async def manual_exp_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    t = update.message.text.strip().lower()
    if t in ("сегодня", "today"):
        context.user_data["exp"]["дата"] = datetime.now().strftime("%d.%m.%Y")
    else:
        dt = parse_date_or_none(update.message.text)
        if dt is None:
            await update.message.reply_text("⚠️ Не понял дату. Введи в виде 08.06.2026 или «сегодня».")
            return E_DATE
        context.user_data["exp"]["дата"] = date_to_str(dt)
    await update.message.reply_text("💵 Сумма расхода, тенге?")
    return E_AMOUNT


async def manual_exp_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    amt = parse_num(update.message.text)
    if not amt:
        await update.message.reply_text("⚠️ Введи сумму числом, например 150000.")
        return E_AMOUNT
    context.user_data["exp"]["сумма"] = amt
    kb = ReplyKeyboardMarkup([["Производство"], ["Администрация"], ["Капзатраты"], ["❌ Отмена"]],
                             resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text("📂 Группа расхода?", reply_markup=kb)
    return E_GROUP


async def manual_exp_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    grp = update.message.text.strip()
    if grp not in EXPENSE_GROUPS:
        await update.message.reply_text("⚠️ Выбери кнопкой: Производство, Администрация или Капзатраты.")
        return E_GROUP
    context.user_data["exp"]["группа"] = grp
    kb = ReplyKeyboardMarkup([[c] for c in EXPENSE_GROUPS[grp]] + [["⬅️ Назад", "❌ Отмена"]],
                             resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text("🏷 Категория?", reply_markup=kb)
    return E_CATEGORY


async def manual_exp_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cat = update.message.text.strip()
    if "назад" in cat.lower():
        kb = ReplyKeyboardMarkup([["Производство"], ["Администрация"], ["Капзатраты"], ["❌ Отмена"]],
                                 resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text("📂 Группа расхода?", reply_markup=kb)
        return E_GROUP
    grp = context.user_data["exp"].get("группа", "")
    if cat not in EXPENSE_GROUPS.get(grp, []):
        await update.message.reply_text("⚠️ Выбери категорию кнопкой из списка. Или ⬅️ Назад / ❌ Отмена.")
        return E_CATEGORY
    context.user_data["exp"]["категория"] = cat
    if context.user_data["exp"].get("источник"):
        return await _ask_note(update, context)
    kb = ReplyKeyboardMarkup([["Евразийский", "БЦК"], ["Касса 1", "Касса 2"], ["Неденежный"], ["⬅️ Назад", "❌ Отмена"]],
                             resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text("🏦 Источник (откуда оплата)?", reply_markup=kb)
    return E_SOURCE


async def manual_exp_source(update: Update, context: ContextTypes.DEFAULT_TYPE):
    src = update.message.text.strip()
    if "назад" in src.lower():
        grp = context.user_data["exp"].get("группа", "")
        kb = ReplyKeyboardMarkup([[c] for c in EXPENSE_GROUPS.get(grp, [])] + [["⬅️ Назад", "❌ Отмена"]],
                                 resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text("🏷 Категория?", reply_markup=kb)
        return E_CATEGORY
    if src not in EXPENSE_SOURCES:
        await update.message.reply_text("⚠️ Выбери источник кнопкой. Или ⬅️ Назад / ❌ Отмена.")
        return E_SOURCE
    context.user_data["exp"]["источник"] = src
    if context.user_data["exp"].get("_photo"):
        return await _ask_note(update, context)
    await update.message.reply_text("🏢 Контрагент (кому платим)? Или «-», если не нужно.")
    return E_CONTRAGENT


async def manual_exp_contragent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    t = update.message.text.strip()
    name = "" if t in ("-", "—") else t
    nxt = await _supplier_resolve(update, context, name)
    if nxt is not None:
        return nxt
    await update.message.reply_text("📝 Примечание? Или «-», если нет.")
    return E_NOTE


async def _ask_note(update, context):
    """Шаг «за что платим». Если с фото распознано назначение — предлагаем оставить его."""
    rec = context.user_data["exp"].get("примечание", "")
    if rec:
        kb = ReplyKeyboardMarkup([["✅ Оставить как есть"], ["⬅️ Назад", "❌ Отмена"]],
                                 resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text(
            f"📝 За что платим? Распознал с документа: «{rec}»\n"
            "Нажми «✅ Оставить как есть» или напиши своё (или «-», если не нужно).",
            reply_markup=kb)
    else:
        kb = ReplyKeyboardMarkup([["⬅️ Назад", "❌ Отмена"]],
                                 resize_keyboard=True, one_time_keyboard=True)
        await update.message.reply_text(
            "📝 За что платим? Напиши примечание (или «-», если не нужно).", reply_markup=kb)
    return E_NOTE


async def manual_exp_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    t = update.message.text.strip()
    low = t.lower()
    if "назад" in low:
        if context.user_data["exp"].get("_photo"):
            grp = context.user_data["exp"].get("группа", "")
            kb = ReplyKeyboardMarkup([[c] for c in EXPENSE_GROUPS.get(grp, [])] + [["⬅️ Назад", "❌ Отмена"]],
                                     resize_keyboard=True, one_time_keyboard=True)
            await update.message.reply_text("🏷 Категория?", reply_markup=kb)
            return E_CATEGORY
        await update.message.reply_text("🏢 Контрагент (кому платим)? Или «-», если не нужно.")
        return E_CONTRAGENT
    if "оставить" not in low:
        context.user_data["exp"]["примечание"] = "" if t in ("-", "—") else t
    kb = ReplyKeyboardMarkup([["✅ Да, сохранить"], ["⬅️ Назад", "❌ Отмена"]],
                             resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text(_exp_summary(context.user_data["exp"], saved=False),
                                    reply_markup=kb, parse_mode="Markdown")
    return E_CONFIRM


async def manual_exp_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ans = update.message.text.strip().lower()
    if "назад" in ans:
        return await _ask_note(update, context)
    if "да" in ans or "сохран" in ans:
        d = context.user_data["exp"]
        save_expense(d)
        schedule_refresh()
        await update.message.reply_text(_exp_summary(d, saved=True), parse_mode="Markdown")
        context.user_data.pop("exp", None)
        return ConversationHandler.END
    if "отмена" in ans or "нет" in ans:
        context.user_data.pop("exp", None)
        await update.message.reply_text("❌ Отменено, расход не сохранён.")
        return ConversationHandler.END
    await update.message.reply_text("Нажми «✅ Да, сохранить» или «❌ Отмена».")
    return E_CONFIRM


async def recognize_expense(image_bytes: bytes):
    """Распознаёт расходный документ (счёт/чек/накладную на покупку)."""
    image_b64 = base64.b64encode(image_bytes).decode()
    prompt = (
        "Это фото счёта, чека, платёжного поручения или накладной на РАСХОД (компания оплачивает товар или услугу).\n"
        "Верни ТОЛЬКО JSON без markdown:\n"
        '{"дата":"дд.мм.гггг","контрагент":"кому платим (поставщик/исполнитель)","сумма":0,'
        '"банк":"банк плательщика, если виден","назначение":"за что платим — назначение платежа '
        'или наименование товара/услуги, коротко своими словами"}\n'
        "Сумма — итог к оплате. Контрагент — ПОЛУЧАТЕЛЬ платежа. Банк — банк ПЛАТЕЛЬЩИКА (наш счёт, откуда деньги), если виден. "
        "Назначение — короткая суть (например «электроэнергия за май», «запчасти к станку»). Если что-то не читается — 0 или пусто."
    )
    return await call_claude(image_b64, prompt)


async def manual_exp_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Фото счёта/чека внутри «Ввод расхода»: распознаём сумму/контрагента/дату."""
    await update.message.reply_text("📸 Получил фото, распознаю расход...")
    try:
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        image_bytes = bytes(await file.download_as_bytearray())
        data = await recognize_expense(image_bytes)
        logger.info(f"Expense photo data: {data}")
    except Exception as e:
        logger.error(f"manual_exp_photo error: {e}")
        await update.message.reply_text("❌ Не смог распознать фото. Введи дату вручную (дд.мм.гггг) или пришли фото чётче.")
        return E_DATE
    return await _exp_from_photo_data(update, context, data)


async def _exp_from_photo_data(update, context, data):
    """Заполняет черновик расхода из распознанного фото и ведёт дальше по шагам.
    Используется и внутри «💸 Ввод расхода», и когда фото-расход пришёл без старта."""
    dt = parse_date_or_none(data.get("дата", ""))
    context.user_data["exp"] = {
        "_photo": True,
        "дата": date_to_str(dt) if dt else datetime.now().strftime("%d.%m.%Y"),
        "сумма": parse_num(data.get("сумма", 0)),
        "контрагент": str(data.get("контрагент", "")).strip(),
        "примечание": str(data.get("назначение", "")).strip()[:120],
    }
    bank_raw = str(data.get("банк", "")).lower()
    if "евраз" in bank_raw:
        context.user_data["exp"]["источник"] = "Евразийский"
    elif "центркредит" in bank_raw or "бцк" in bank_raw or "centercredit" in bank_raw:
        context.user_data["exp"]["источник"] = "БЦК"
    d = context.user_data["exp"]
    await update.message.reply_text(
        f"💸 Распознал:\n💵 Сумма: {d['сумма']} тнг\n🏢 Контрагент: {d['контрагент'] or '—'}\n📅 Дата: {d['дата']}"
    )
    nxt = await _supplier_resolve(update, context, context.user_data["exp"].get("контрагент", ""))
    if nxt is not None:
        return nxt
    kb = ReplyKeyboardMarkup([["Производство"], ["Администрация"], ["Капзатраты"], ["❌ Отмена"]],
                             resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text("Выбери группу расхода:", reply_markup=kb)
    return E_GROUP


def _build_svod():
    """Пересобирает лист «Свод расходов»: категории × месяцы, суммы из журнала «Расходы» (готовые числа, не формулы)."""
    book = get_spreadsheet()
    sums = {}
    try:
        jrows = book.worksheet("Расходы").get_all_values()[1:]
    except gspread.WorksheetNotFound:
        jrows = []
    year = datetime.now().year
    for row in jrows:
        if len(row) < 4:
            continue
        s = parse_num(row[1])
        if not s:
            continue
        dt = parse_date_or_none(row[0])
        if not dt or dt.year != year:
            continue
        cat = str(row[3]).strip()
        sums.setdefault(cat, {})
        sums[cat][dt.month] = sums[cat].get(dt.month, 0) + s
    title = "Свод расходов"
    try:
        ws = book.worksheet(title)
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = book.add_worksheet(title=title, rows=60, cols=15)
    months = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн", "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]
    data = [["Категория"] + months + ["Итого"]]
    for grp in ("Производство", "Администрация", "Капзатраты"):
        data.append([grp.upper()] + [""] * 13)
        for cat in EXPENSE_GROUPS[grp]:
            row_vals = [cat]
            total = 0
            for mi in range(1, 13):
                v = sums.get(cat, {}).get(mi, 0)
                row_vals.append(round(v))
                total += v
            row_vals.append(round(total))
            data.append(row_vals)
    ws.update(range_name="A1", values=data, value_input_option="USER_ENTERED")
    return title, year


def _build_sebes():
    """Лист «Себестоимость (авто)»: по месяцам — производство, расходы по категориям, полная себестоимость, выручка, прибыль, себестоимость 1 кг."""
    book = get_spreadsheet()
    year = datetime.now().year
    tyres = {m: 0.0 for m in range(1, 13)}
    crumb = {m: 0.0 for m in range(1, 13)}
    metal = {m: 0.0 for m in range(1, 13)}
    rev = {m: 0.0 for m in range(1, 13)}
    rev_novat = {m: 0.0 for m in range(1, 13)}
    cat_sums = {}
    try:
        prows = book.worksheet("Производство").get_all_values()[1:]
    except gspread.WorksheetNotFound:
        prows = []
    for row in prows:
        if len(row) < 12:
            row = row + [""] * (12 - len(row))
        if "начальн" in str(row[1]).strip().lower():
            continue
        dt = parse_date_or_none(row[0])
        if not dt or dt.year != year:
            continue
        m = dt.month
        tyres[m] += parse_num(row[2]); crumb[m] += parse_num(row[10]); metal[m] += parse_num(row[11])
    try:
        srows = book.worksheet("Реализация").get_all_values()[1:]
    except gspread.WorksheetNotFound:
        srows = []
    for row in srows:
        if len(row) < 8:
            row = row + [""] * (8 - len(row))
        dt = parse_date_or_none(row[0])
        if not dt or dt.year != year:
            continue
        m = dt.month
        v = parse_num(row[6]); rev[m] += v; rev_novat[m] += v - parse_num(row[7])
    try:
        jrows = book.worksheet("Расходы").get_all_values()[1:]
    except gspread.WorksheetNotFound:
        jrows = []
    for row in jrows:
        if len(row) < 4:
            continue
        s = parse_num(row[1])
        if not s:
            continue
        dt = parse_date_or_none(row[0])
        if not dt or dt.year != year:
            continue
        cat = str(row[3]).strip()
        cat_sums.setdefault(cat, {})
        cat_sums[cat][dt.month] = cat_sums[cat].get(dt.month, 0) + s

    prodcats = EXPENSE_GROUPS["Производство"]
    admcats = EXPENSE_GROUPS["Администрация"]
    def cs(c, m):
        return cat_sums.get(c, {}).get(m, 0)
    def prodc(m):
        return sum(cs(c, m) for c in prodcats)
    def admc(m):
        return sum(cs(c, m) for c in admcats)
    def full(m):
        return prodc(m) + admc(m)
    months = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн", "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]
    def vals(fn):
        return [round(fn(m)) for m in range(1, 13)]
    data = [["Показатель"] + months]
    data.append(["ПРОИЗВОДСТВО, кг"] + [""] * 12)
    data.append(["Переработано шин"] + vals(lambda m: tyres[m]))
    data.append(["Произведено крошки"] + vals(lambda m: crumb[m]))
    data.append(["Металлокорд"] + vals(lambda m: metal[m]))
    data.append(["Выход крошки, %"] + [round(crumb[m] / tyres[m] * 100, 1) if tyres[m] else 0 for m in range(1, 13)])
    data.append(["ПРОИЗВОДСТВЕННАЯ СЕБЕСТОИМОСТЬ"] + [""] * 12)
    for c in prodcats:
        data.append([c] + vals(lambda m, c=c: cs(c, m)))
    data.append(["ИТОГО производственная"] + vals(prodc))
    data.append(["АДМИНИСТРАТИВНЫЕ РАСХОДЫ"] + [""] * 12)
    for c in admcats:
        data.append([c] + vals(lambda m, c=c: cs(c, m)))
    data.append(["ИТОГО административные"] + vals(admc))
    data.append(["ПОЛНАЯ СЕБЕСТОИМОСТЬ"] + vals(full))
    data.append(["СЕБЕСТОИМОСТЬ 1 кг, тенге"] + [""] * 12)
    data.append(["Производственная 1 кг"] + [round(prodc(m) / crumb[m], 2) if crumb[m] else 0 for m in range(1, 13)])
    data.append(["Полная 1 кг"] + [round(full(m) / crumb[m], 2) if crumb[m] else 0 for m in range(1, 13)])
    data.append(["ВЫРУЧКА И ПРИБЫЛЬ"] + [""] * 12)
    data.append(["Выручка с НДС"] + vals(lambda m: rev[m]))
    data.append(["Выручка без НДС"] + vals(lambda m: rev_novat[m]))
    data.append(["Прибыль (без НДС − полная себест.)"] + vals(lambda m: rev_novat[m] - full(m)))
    title = "Себестоимость (авто)"
    try:
        ws = book.worksheet(title); ws.clear()
    except gspread.WorksheetNotFound:
        ws = book.add_worksheet(title=title, rows=45, cols=14)
    ws.update(range_name="A1", values=data, value_input_option="USER_ENTERED")
    return title


async def cmd_svod(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Собрать/обновить лист «Свод расходов». Только для разрешённых."""
    if not is_allowed(update):
        await update.message.reply_text("⛔ Нет доступа.")
        return
    await update.message.reply_text("📊 Собираю «Свод расходов»...")
    try:
        title, year = await asyncio.to_thread(_build_svod)
        title2 = await asyncio.to_thread(_build_sebes)
    except Exception as e:
        logger.error(f"svod error: {e}")
        await update.message.reply_text(f"❌ Ошибка при сборке свода: {e}")
        return
    await update.message.reply_text(
        f"✅ Готово (год {year}): листы «{title}» и «{title2}». "
        f"«Свод расходов» — расходы по категориям; «Себестоимость (авто)» — производство, расходы, полная себестоимость, выручка, прибыль и себестоимость 1 кг по месяцам."
    )


# --- Запись себестоимости в файл экономиста («Калькуляция себестоимости») ---
SEBES_FILE_ID = "1ZKYCFVKrb0l-mzYQ0gtiHKOJbTNHd9TcJ9hN9i5RiFk"
KALK_SHEET = "Калькуляция себестоимости"
RU_MONTHS_FULL = {1: "январь", 2: "февраль", 3: "март", 4: "апрель", 5: "май", 6: "июнь",
                  7: "июль", 8: "август", 9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"}
# строка «Калькуляции» -> категория журнала (или спец-ключ производства)
KALK_ROWS = {
    4: "_шины", 5: "_крошка", 6: "_металл",
    9: "Материалы и запчасти", 10: "Упаковка (мешки)", 11: "Электроэнергия", 12: "ГСМ",
    13: "ФОТ производственный", 14: "Транспорт и логистика", 15: "Лизинг (вознаграждение)",
    16: "Амортизация производства",
    19: "ФОТ административный", 20: "Услуги банка", 21: "Налоги", 22: "Аудит", 23: "Страхование",
    24: "Прочие административные", 25: "Амортизация административная",
    32: "Лизинг (основной долг)", 33: "Основные средства",
}


def _col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _ensure_kalk_column(book, ws, month):
    """Находит столбец месяца в «Калькуляции» или создаёт его (перед «Итого», с форматом соседа).
    Возвращает (col_1based, created)."""
    target = RU_MONTHS_FULL[month]
    header = ws.row_values(2)
    for i, h in enumerate(header, start=1):
        if target in str(h).strip().lower():
            return i, False
    itog_idx = None
    for i, h in enumerate(header, start=1):
        if "итог" in str(h).strip().lower():
            itog_idx = i
            break
    if itog_idx is None:
        itog_idx = len(header) + 1
    book.batch_update({"requests": [
        {"insertDimension": {"range": {"sheetId": ws.id, "dimension": "COLUMNS",
                                       "startIndex": itog_idx - 1, "endIndex": itog_idx},
                             "inheritFromBefore": True}}
    ]})
    ws.update_cell(2, itog_idx, target.capitalize())
    return itog_idx, True


def _fill_kalk(month):
    """Заполняет (и при необходимости создаёт) столбец месяца в «Калькуляции». Возвращает (col, crumb, created)."""
    skl = get_spreadsheet()
    year = datetime.now().year
    tyres = crumb = metal = 0.0
    for row in skl.worksheet("Производство").get_all_values()[1:]:
        if len(row) < 12:
            row = row + [""] * (12 - len(row))
        if "начальн" in str(row[1]).strip().lower():
            continue
        dt = parse_date_or_none(row[0])
        if not dt or dt.year != year or dt.month != month:
            continue
        tyres += parse_num(row[2]); crumb += parse_num(row[10]); metal += parse_num(row[11])
    cat = {}
    try:
        for row in skl.worksheet("Расходы").get_all_values()[1:]:
            if len(row) < 4:
                continue
            s = parse_num(row[1])
            if not s:
                continue
            dt = parse_date_or_none(row[0])
            if not dt or dt.year != year or dt.month != month:
                continue
            c = str(row[3]).strip()
            cat[c] = cat.get(c, 0) + s
    except gspread.WorksheetNotFound:
        pass
    # налоги на ФОТ — в себестоимость, пропорционально зарплате произв./адм. (решение 12.06)
    tax_total = sum(v for k, v in cat.items() if str(k).startswith("Налог ФОТ"))
    if tax_total:
        fot_p = cat.get("ФОТ производственный", 0)
        fot_a = cat.get("ФОТ административный", 0)
        share_p = fot_p / (fot_p + fot_a) if (fot_p + fot_a) else 0.5
        cat["ФОТ производственный"] = fot_p + tax_total * share_p
        cat["ФОТ административный"] = fot_a + tax_total * (1 - share_p)
    book = _get_client().open_by_key(SEBES_FILE_ID)
    ws = book.worksheet(KALK_SHEET)
    col_idx, created = _ensure_kalk_column(book, ws, month)
    col = _col_letter(col_idx)
    vals = {"_шины": tyres, "_крошка": crumb, "_металл": metal}
    updates = []
    for r, key in KALK_ROWS.items():
        v = vals[key] if key in vals else cat.get(key, 0)
        updates.append({"range": col + str(r), "values": [[round(v)]]})
    # итоги и себестоимость пишем при КАЖДОМ запуске (не только при создании столбца),
    # иначе при авто-обновлении категории обновляются, а итоги отстают
    prod_t = sum(cat.get(k, 0) for k in ["Материалы и запчасти", "Упаковка (мешки)", "Электроэнергия",
                 "ГСМ", "ФОТ производственный", "Транспорт и логистика", "Лизинг (вознаграждение)",
                 "Амортизация производства"])
    adm_t = sum(cat.get(k, 0) for k in ["ФОТ административный", "Услуги банка", "Налоги", "Аудит",
                "Страхование", "Прочие административные", "Амортизация административная"])
    full = prod_t + adm_t
    kap = cat.get("Лизинг (основной долг)", 0) + cat.get("Основные средства", 0)
    extra = {7: round(crumb / tyres, 3) if tyres else 0, 17: round(prod_t), 26: round(adm_t),
             27: round(full), 29: round(prod_t / crumb, 2) if crumb else 0,
             30: round(full / crumb, 2) if crumb else 0, 35: round(kap)}
    for r, v in extra.items():
        updates.append({"range": col + str(r), "values": [[v]]})
    ws.batch_update(updates, value_input_option="USER_ENTERED")
    return col, crumb, created


async def cmd_kalk(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Заполнить текущий месяц в «Калькуляции» файла экономиста."""
    if not is_allowed(update):
        await update.message.reply_text("⛔ Нет доступа.")
        return
    month = datetime.now().month
    await update.message.reply_text(f"📊 Заполняю «Калькуляцию» за {RU_MONTHS_FULL[month]}...")
    try:
        res = await asyncio.to_thread(_fill_kalk, month)
    except Exception as e:
        logger.error(f"kalk error: {e}")
        await update.message.reply_text(f"❌ Ошибка: {e}. Проверь, что файл открыт боту на «Редактор».")
        return
    col, crumb, created = res
    note = " (создал столбец автоматически)" if created else ""
    await update.message.reply_text(
        f"✅ Столбец «{RU_MONTHS_FULL[month]}»{note} в «Калькуляции» заполнен: производство + расходы. "
        f"Произведено крошки: {round(crumb)} кг. Итоги и себестоимость 1 кг посчитаны.")


# ---------- /fin: автозаполнение месячной вкладки финотчёта (v1) ----------
# Бот заполняет ТОЛЬКО свои блоки: безнал-поступления (ERG / Юр.лица, по банкам),
# 5 групп безнал-расходов и наличные расходы Касса 1 / Касса 2.
# Блоки бухгалтера (ФОТ с разбивкой, налоги, лизинг, банк, снятие, переводы группы,
# все наличные поступления) бот НЕ трогает. Строки ищутся по меткам, не по номерам.

FIN_BANKS = ("Евразийский", "БЦК")
# категория журнала «Расходы» -> группа безнал-расхода в финотчёте (метка строки)
FIN_GROUP_OF_CAT = {
    "Основные средства": "основные средства",
    "Прочие капзатраты": "основные средства",
    "Материалы и запчасти": "материалы и запасные части",
    "ГСМ": "гсм",
    "Транспорт и логистика": "транспортные услуги",
    "Упаковка (мешки)": "прочие расходы",
    "Электроэнергия": "прочие расходы",
    "Прочие административные": "прочие расходы",
    "Страхование": "прочие расходы",
    "Аудит": "прочие расходы",
    # лизинг АО «ФРП» — решение от 12.06: вводится через бот по платёжкам, не вручную
    "Лизинг (основной долг)": "ао «фонд",
    "Лизинг (вознаграждение)": "ао «фонд",
}
# для лизинга подстроки называются по сути платежа, а не по контрагенту
FIN_LEAF_NAME = {
    "Лизинг (основной долг)": "основной долг",
    "Лизинг (вознаграждение)": "вознаграждение",
}
FIN_GROUPS = ["основные средства", "материалы и запасные части", "гсм",
              "транспортные услуги", "ао «фонд", "прочие расходы"]


def _fin_norm(s):
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _bank_of_sale_note(note):
    """Банк из примечания продажи ('банк: БЦК'). Старые записи без банка -> Евразийский."""
    return "БЦК" if "бцк" in str(note).lower() else "Евразийский"


def _is_erg_buyer(buyer):
    b = str(buyer).lower()
    return "ссгпо" in b or "erg" in b


def _fin_collect(month):
    """Данные месяца из склада: безнал-поступления (продажи), безнал- и нал-расходы."""
    skl = get_spreadsheet()
    year = datetime.now().year
    inc = {"erg": {b: 0.0 for b in FIN_BANKS}, "jur": {b: 0.0 for b in FIN_BANKS}}
    salary = {b: 0.0 for b in FIN_BANKS}
    fot_tax = {}  # вид налога ФОТ -> {банк: сумма}
    for row in skl.worksheet(SHEET_SALES).get_all_values()[1:]:
        if len(row) < 9:
            row = row + [""] * (9 - len(row))
        dt = parse_date_or_none(row[0])
        if not dt or dt.year != year or dt.month != month:
            continue
        if "без" not in str(row[2]).lower():
            continue
        s = parse_num(row[6])
        bank = _bank_of_sale_note(row[8])
        key = "erg" if _is_erg_buyer(row[1]) else "jur"
        inc[key][bank] += s
    cashless = {}  # группа -> контрагент -> {банк: сумма}
    cash = {"Касса 1": {}, "Касса 2": {}}  # касса -> метка -> сумма
    try:
        rows = skl.worksheet(SHEET_EXPENSES).get_all_values()[1:]
    except gspread.WorksheetNotFound:
        rows = []
    for row in rows:
        if len(row) < 8:
            row = row + [""] * (8 - len(row))
        dt = parse_date_or_none(row[0])
        if not dt or dt.year != year or dt.month != month:
            continue
        s = parse_num(row[1])
        if not s:
            continue
        cat, src, agent = str(row[3]).strip(), str(row[5]).strip(), str(row[6]).strip()
        if src in FIN_BANKS:
            if cat in ("ФОТ производственный", "ФОТ административный"):
                salary[src] += s  # зарплата на карты — строка внутри блока ФОТ
                continue
            if cat.startswith("Налог ФОТ"):
                vid = cat.split(":", 1)[1].strip() if ":" in cat else cat
                d2 = fot_tax.setdefault(vid, {b: 0.0 for b in FIN_BANKS})
                d2[src] += s
                continue
            grp = FIN_GROUP_OF_CAT.get(cat)
            if not grp:
                continue  # налоги, банк — заполняет бухгалтер
            name = FIN_LEAF_NAME.get(cat) or agent or cat
            d = cashless.setdefault(grp, {}).setdefault(name, {b: 0.0 for b in FIN_BANKS})
            d[src] += s
        elif src in ("Касса 1", "Касса 2"):
            name = cat + (f" ({agent})" if agent else "")
            cash[src][name] = cash[src].get(name, 0) + s
    return inc, cashless, cash, salary, fot_tax


def _fin_find(grid, needle, start=0, end=None):
    """Индекс строки (0-based), метка в колонке A начинается с needle."""
    e = len(grid) if end is None else end
    for i in range(start, e):
        a = _fin_norm(grid[i][0] if grid[i] else "")
        if a.startswith(needle):
            return i
    return None


def _fin_leaf_range(grid, gi):
    """Диапазон строк-листьев группы из формулы B '=SUM(B22:B24)'. None — группа без листьев."""
    b = str(grid[gi][1]) if len(grid[gi]) > 1 else ""
    m = re.search(r"SUM\(B(\d+)(?::B(\d+))?\)", b, re.I)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2) or m.group(1))


def _fin_insert_rows(book, ws, after_row, count):
    """Вставляет count строк ПОСЛЕ 1-based строки after_row (внутри диапазона, формат сверху)."""
    book.batch_update({"requests": [{"insertDimension": {
        "range": {"sheetId": ws.id, "dimension": "ROWS",
                  "startIndex": after_row, "endIndex": after_row + count},
        "inheritFromBefore": True}}]})


def _fill_fin(month):
    """Заполняет вкладку месяца в финотчёте. Возвращает текст-сводку."""
    inc, cashless, cash, salary, fot_tax = _fin_collect(month)
    book = _get_client().open_by_key(SEBES_FILE_ID)
    name = RU_MONTHS_FULL[month]
    year = datetime.now().year
    created = False
    try:
        ws = book.worksheet(name)
    except gspread.WorksheetNotFound:
        prev = None
        for m in range(month - 1, 0, -1):
            try:
                prev = book.worksheet(RU_MONTHS_FULL[m])
                break
            except gspread.WorksheetNotFound:
                continue
        if prev is None:
            raise RuntimeError("не нашёл вкладку прошлого месяца как шаблон")
        ws = prev.duplicate(new_sheet_name=name, insert_sheet_index=prev.index + 1)
        created = True

    grid = ws.get(value_render_option="FORMULA")
    zeroed = False
    head = _fin_norm(grid[0][0] if grid and grid[0] else "")
    if name not in head:
        # первый запуск месяца: обнуляем все числа (формулы не трогаем), правим заголовок
        ups = [{"range": "A1",
                "values": [[f'Отчёт ТОО "Едiл и компания" за {name.upper()} {year} г.']]}]
        for i, row in enumerate(grid[3:], start=4):
            for j in (1, 2, 3):
                v = row[j] if len(row) > j else ""
                if v == "" or str(v).startswith("="):
                    continue
                ups.append({"range": "ABCD"[j] + str(i), "values": [[0]]})
        ws.batch_update(ups, value_input_option="USER_ENTERED")
        zeroed = True
        grid = ws.get(value_render_option="FORMULA")

    i_rash = _fin_find(grid, "безналичный расход")
    i_nalpost = _fin_find(grid, "наличные поступления")
    i_nalrash = _fin_find(grid, "наличные расходы")
    if None in (i_rash, i_nalpost, i_nalrash):
        raise RuntimeError("не нашёл разделы вкладки (БЕЗНАЛИЧНЫЙ РАСХОД / НАЛИЧНЫЕ ...)")

    # 1) безнал-поступления от продаж: ERG и Юр.лица, по банкам
    ups = []
    for needle, key in (("erg", "erg"), ("юр", "jur")):
        i = _fin_find(grid, needle, 0, i_rash)
        if i is not None:
            ups.append({"range": f"B{i + 1}", "values": [[round(inc[key]["Евразийский"], 2)]]})
            ups.append({"range": f"C{i + 1}", "values": [[round(inc[key]["БЦК"], 2)]]})
    if ups:
        ws.batch_update(ups, value_input_option="USER_ENTERED")

    # 1.5) блок ФОТ: зарплата + оплаченные налоги ФОТ (по ведомости через бот).
    # Исполнительные листы — бухгалтер, не трогаем.
    if salary["Евразийский"] or salary["БЦК"] or fot_tax:
        gi_fot = _fin_find(grid, "фот", i_rash, i_nalpost)
        rng_fot = _fin_leaf_range(grid, gi_fot) if gi_fot is not None else None
        if rng_fot:
            ups2 = []
            if salary["Евразийский"] or salary["БЦК"]:
                zi = _fin_find(grid, "заработная плата", rng_fot[0] - 1, rng_fot[1])
                if zi is not None:
                    ups2.append({"range": f"B{zi + 1}", "values": [[round(salary["Евразийский"], 2)]]})
                    ups2.append({"range": f"C{zi + 1}", "values": [[round(salary["БЦК"], 2)]]})
            # ОПВР раньше ОПВ, чтобы префиксы не перепутались; занятые строки помечаем
            used_rows = set()
            for vid in ("ОПВР", "ОПВ", "ИПН", "Социальный налог", "Социальные отчисления",
                        "ВОСМС", "ОСМС"):
                v = fot_tax.get(vid)
                if not v:
                    continue
                needle = "налог фот: " + vid.lower()
                ti = None
                for i in range(rng_fot[0] - 1, rng_fot[1]):
                    if i in used_rows:
                        continue
                    a = _fin_norm(grid[i][0] if grid[i] else "")
                    if a.startswith(needle):
                        ti = i
                        break
                if ti is None:
                    continue
                used_rows.add(ti)
                ups2.append({"range": f"B{ti + 1}", "values": [[round(v["Евразийский"], 2)]]})
                ups2.append({"range": f"C{ti + 1}", "values": [[round(v["БЦК"], 2)]]})
            if ups2:
                ws.batch_update(ups2, value_input_option="USER_ENTERED")

    # 2) наличные расходы: Касса 2, затем Касса 1 (снизу вверх, чтобы вставки не сдвигали)
    for kassa in ("Касса 2", "Касса 1"):
        h = _fin_find(grid, _fin_norm(kassa), i_nalrash)
        if h is None:
            continue
        it = _fin_find(grid, "итого касса", h + 1)
        if it is None:
            continue
        r1, r2 = h + 2, it  # 1-based листья
        items = sorted(cash[kassa].items(), key=lambda kv: -kv[1])
        need = max(len(items), 1)
        ins = need - (r2 - r1 + 1)
        if ins > 0:
            _fin_insert_rows(book, ws, r1, ins)
            r2 += ins
        data = []
        for k in range(r2 - r1 + 1):
            r = r1 + k
            if k < len(items):
                nm, v = items[k]
                data.append({"range": f"A{r}", "values": [["    " + nm]]})
                data.append({"range": f"D{r}", "values": [[round(v, 2)]]})
            else:
                data.append({"range": f"A{r}", "values": [["    —"]]})
                data.append({"range": f"D{r}", "values": [[0]]})
        data.append({"range": f"D{r2 + 1}", "values": [[f"=SUM(D{r1}:D{r2})"]]})
        ws.batch_update(data, value_input_option="USER_ENTERED")

    # 3) безнал-расходы: 5 групп бота, снизу вверх
    for gname in reversed(FIN_GROUPS):
        gi = _fin_find(grid, gname, i_rash, i_nalpost)
        if gi is None:
            continue
        items = sorted(cashless.get(gname, {}).items(),
                       key=lambda kv: -(kv[1]["Евразийский"] + kv[1]["БЦК"]))
        rng = _fin_leaf_range(grid, gi)
        if rng is None and not items:
            # строка-группа без подстрок и без данных — просто нули
            ws.batch_update([{"range": f"B{gi + 1}", "values": [[0]]},
                             {"range": f"C{gi + 1}", "values": [[0]]}],
                            value_input_option="USER_ENTERED")
            continue
        if rng is None:
            # строка-группа без подстрок (как ГСМ): превращаем в группу с контрагентами,
            # чтобы было видно, кому платили — как в «Прочих расходах»
            _fin_insert_rows(book, ws, gi + 1, len(items))
            r1, r2 = gi + 2, gi + 1 + len(items)
        else:
            r1, r2 = rng
            need = max(len(items), 1)
            ins = need - (r2 - r1 + 1)
            if ins > 0:
                _fin_insert_rows(book, ws, r1, ins)
                r2 += ins
        data = []
        for k in range(r2 - r1 + 1):
            r = r1 + k
            if k < len(items):
                nm, v = items[k]
                data.append({"range": f"A{r}", "values": [["    " + nm]]})
                data.append({"range": f"B{r}", "values": [[round(v["Евразийский"], 2)]]})
                data.append({"range": f"C{r}", "values": [[round(v["БЦК"], 2)]]})
            else:
                data.append({"range": f"A{r}", "values": [["    —"]]})
                data.append({"range": f"B{r}", "values": [[0]]})
                data.append({"range": f"C{r}", "values": [[0]]})
            data.append({"range": f"D{r}", "values": [[f"=SUM(B{r}:C{r})"]]})
        data.append({"range": f"B{gi + 1}", "values": [[f"=SUM(B{r1}:B{r2})"]]})
        data.append({"range": f"C{gi + 1}", "values": [[f"=SUM(C{r1}:C{r2})"]]})
        data.append({"range": f"D{gi + 1}", "values": [[f"=SUM(B{gi + 1}:C{gi + 1})"]]})
        ws.batch_update(data, value_input_option="USER_ENTERED")

    t_inc = sum(inc[k][b] for k in inc for b in FIN_BANKS)
    t_out = sum(v[b] for g in cashless.values() for v in g.values() for b in FIN_BANKS)
    t_k1 = sum(cash["Касса 1"].values())
    t_k2 = sum(cash["Касса 2"].values())
    notes = []
    if created:
        notes.append("вкладка создана по шаблону прошлого месяца")
    if zeroed:
        notes.append("старые цифры обнулены")
    extra = (" (" + ", ".join(notes) + ")") if notes else ""
    t_sal = sum(salary.values())
    t_tax = sum(sum(v.values()) for v in fot_tax.values())
    return (f"✅ Финотчёт: вкладка «{name}» заполнена{extra}.\n\n"
            f"💰 Безнал-поступления (продажи): {round(t_inc)} тнг\n"
            f"💸 Безнал-расходы (группы бота): {round(t_out)} тнг\n"
            f"👥 Зарплата на карты (ФОТ): {round(t_sal)} тнг\n"
            f"🧾 Налоги ФОТ (оплачено): {round(t_tax)} тнг\n"
            f"💵 Нал-расходы: Касса 1 — {round(t_k1)}, Касса 2 — {round(t_k2)} тнг\n\n"
            f"Блоки бухгалтера (исп. листы, прочие налоги, переводы, нал-поступления) не тронуты.")


async def cmd_fin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Заполнить вкладку текущего месяца в финотчёте экономиста."""
    if not is_allowed(update):
        await update.message.reply_text("⛔ Нет доступа.")
        return
    month = datetime.now().month
    await update.message.reply_text(f"📊 Заполняю финотчёт за {RU_MONTHS_FULL[month]}...")
    try:
        summary = await asyncio.to_thread(_fill_fin, month)
    except Exception as e:
        logger.error(f"fin error: {e}")
        await update.message.reply_text(
            f"❌ Ошибка: {e}. Проверь, что файл финотчёта открыт боту на «Редактор».")
        return
    try:
        await asyncio.to_thread(_fill_reports, month)
        summary += "\n\n📈 «Выводы», «Графики» и «Анализ» обновлены по " + RU_MONTHS_FULL[month] + "."
    except Exception as e:
        logger.error(f"reports error: {e}")
        summary += f"\n\n⚠️ Сводные листы не обновились: {e}"
    await update.message.reply_text(summary)


# ---------- сводные листы финотчёта: «Выводы», «Графики», «Анализ и выводы» ----------

VYV_SHEET = "Выводы (5 мес)"   # имя листа НЕ меняем — на него могут быть ссылки
GRAF_SHEET = "Графики"
ANAL_SHEET = "Анализ и выводы"

# строки «Выводов» -> строка на вкладке месяца: (метка в «Выводах», метка месяца, секция)
# секции: 0 = безнал-поступления, 1 = безнал-расход, 2 = наличные и итоги
VYV_MAP = [
    ("erg", "erg", 0),
    ("возвраты", "возвраты", 0),
    ("вознаграждение по депозиту", "вознаграждение по депозиту", 0),
    ("взнос наличными", "взнос наличными", 0),
    ("итого безналичные поступления", "итого безналичные поступления", 0),
    ("переводы между", "переводы между", 0),
    ("внутренний оборот", "внутренний оборот", 0),
    ("итого безнал. поступления", "итого безнал. поступления", 0),
    ("основные средства", "основные средства", 1),
    ("материалы", "материалы", 1),
    ("гсм", "гсм", 1),
    ("фот", "фот", 1),
    ("налоги", "налоги", 1),
    ("транспортные услуги", "транспортные услуги", 1),
    ("лизинг", "ао «фонд", 1),
    ("прочие расходы", "прочие расходы", 1),
    ("банк", "банк", 1),
    ("снятие наличных", "снятие наличных", 1),
    ("итого расход без вн.оборота", "итого расход без вн.оборота", 1),
    ("внутренний оборот", "внутренний оборот", 1),
    ("итого безнал. расход", "итого безналичный расход", 1),
    ("итого наличные поступления", "итого наличные поступления", 2),
    ("всего наличный расход", "всего наличный расход", 2),
    ("всего поступлений", "всего поступлений", 2),
    ("всего расходов", "всего расходов", 2),
]


def _mln(v):
    return f"{v / 1e6:.1f}".replace(".", ",")


def _fill_reports(month):
    """Обновляет «Выводы», «Графики», «Анализ и выводы» по вкладке месяца."""
    book = _get_client().open_by_key(SEBES_FILE_ID)
    name = RU_MONTHS_FULL[month]
    year = datetime.now().year
    mws = book.worksheet(name)
    mg = mws.get(value_render_option="FORMULA")
    mi_rash = _fin_find(mg, "безналичный расход")
    mi_nalpost = _fin_find(mg, "наличные поступления")
    if mi_rash is None or mi_nalpost is None:
        raise RuntimeError("вкладка месяца: не нашёл разделы")
    msec = {0: (0, mi_rash), 1: (mi_rash, mi_nalpost), 2: (mi_nalpost, len(mg))}

    # ===== «Выводы»: столбец месяца формулами на вкладку месяца =====
    vyv = book.worksheet(VYV_SHEET)
    vg = vyv.get(value_render_option="FORMULA")
    hv0 = _fin_find(vg, "поступления (безнал")
    hv1 = _fin_find(vg, "расходы (безнал")
    hv2 = _fin_find(vg, "наличные", (hv1 or 0) + 1)
    if None in (hv0, hv1, hv2):
        raise RuntimeError("«Выводы»: не нашёл разделы")
    vsec = {0: (hv0, hv1), 1: (hv1, hv2), 2: (hv2, len(vg))}
    col_idx, _ = _ensure_kalk_column(book, vyv, month)
    col = _col_letter(col_idx)
    itog = _col_letter(col_idx + 1)
    n_mon = col_idx - 1  # количество месяцев в сводке
    ups = []
    vrows = {}
    used = set()
    for v_needle, m_needle, sec in VYV_MAP:
        a, b = vsec[sec]
        vi = _fin_find(vg, v_needle, a, b)
        a2, b2 = msec[sec]
        mi = _fin_find(mg, m_needle, a2, b2)
        if vi is None or mi is None:
            continue
        key = (v_needle, sec)
        if key in used:
            continue
        used.add(key)
        vrows[v_needle] = vi + 1
        ups.append({"range": f"{col}{vi + 1}", "values": [[f"='{name}'!D{mi + 1}"]]})
        ups.append({"range": f"{itog}{vi + 1}", "values": [[f"=SUM(B{vi + 1}:{col}{vi + 1})"]]})
    rf = vrows.get("фот")
    rl = vrows.get("лизинг")
    ri = vrows.get("итого расход без вн.оборота")
    for needle, num in (("доля фот", rf), ("доля лизинга", rl)):
        di = _fin_find(vg, needle)
        if di is not None and num and ri:
            ups.append({"range": f"{col}{di + 1}",
                        "values": [[f"=IFERROR({col}{num}/{col}{ri},0)"]]})
            ups.append({"range": f"{itog}{di + 1}",
                        "values": [[f"=IFERROR({itog}{num}/{itog}{ri},0)"]]})
    ups.append({"range": "A1", "values": [[
        f'ТОО "Едiл и компания" — сводка и выводы за январь–{name} {year} г. (тенге)']]})
    ups.append({"range": f"{itog}2", "values": [[f"Итого {n_mon} мес"]]})
    vyv.batch_update(ups, value_input_option="USER_ENTERED")

    # адреса итоговых строк на вкладке месяца (нужны «Графикам» и «Анализу»)
    m_post = _fin_find(mg, "всего поступлений") + 1
    m_rash = _fin_find(mg, "всего расходов") + 1
    m_oper = _fin_find(mg, "итого расход без вн.оборота", mi_rash) + 1

    # ===== «Графики»: строка месяца + структура расходов =====
    gws = book.worksheet(GRAF_SHEET)
    gg = gws.get(value_render_option="FORMULA")
    months_low = set(RU_MONTHS_FULL.values())
    first = _fin_find(gg, "январь")
    if first is None:
        raise RuntimeError("«Графики»: не нашёл строку «Январь»")
    last = first
    for i in range(first, len(gg)):
        a = _fin_norm(gg[i][0] if gg[i] else "")
        if a in months_low:
            last = i
        elif i > last:
            break
    mrow = _fin_find(gg, name, first, last + 1)
    shift = 0
    if mrow is None:
        _fin_insert_rows(book, gws, last + 1, 1)
        mrow = last + 1
        shift = 1
    ups = [
        {"range": f"A{mrow + 1}", "values": [[name.capitalize()]]},
        {"range": f"B{mrow + 1}", "values": [[f"='{name}'!D{m_post}"]]},
        {"range": f"C{mrow + 1}", "values": [[f"='{name}'!D{m_rash}"]]},
        {"range": f"D{mrow + 1}", "values": [[f"='{name}'!D{m_oper}"]]},
        {"range": "A1", "values": [[
            f'ТОО "Едiл и компания" — графики за январь–{name} {year} г.']]},
    ]
    pie = [("лизинг", "лизинг"), ("фот", "фот"), ("прочие расходы", "прочие расходы"),
           ("транспортные", "транспортные услуги"), ("основные средства", "основные средства"),
           ("материалы", "материалы")]
    for g_needle, v_needle in pie:
        gi = _fin_find(gg, g_needle, last + 1)
        r = vrows.get(v_needle)
        if gi is not None and r:
            ups.append({"range": f"B{gi + 1 + shift}",
                        "values": [[f"='{VYV_SHEET}'!{itog}{r}"]]})
    gi = _fin_find(gg, "прочее", last + 1)
    parts = [vrows.get(k) for k in ("налоги", "снятие наличных", "гсм", "банк")]
    if gi is not None and all(parts):
        f = "+".join(f"'{VYV_SHEET}'!{itog}{p}" for p in parts)
        ups.append({"range": f"B{gi + 1 + shift}", "values": [["=" + f]]})
    hdr = _fin_find(gg, "статья расхода")
    if hdr is not None:
        ups.append({"range": f"B{hdr + 1 + shift}", "values": [[f"Сумма за {n_mon} мес"]]})
    gws.batch_update(ups, value_input_option="USER_ENTERED")

    # ===== «Анализ и выводы»: строка месяца, ИТОГО, максимумы/минимумы, структура =====
    aws = book.worksheet(ANAL_SHEET)
    ag = aws.get(value_render_option="FORMULA")
    afirst = _fin_find(ag, "январь")
    aitog = _fin_find(ag, "итого", afirst)
    if afirst is None or aitog is None:
        raise RuntimeError("«Анализ»: не нашёл таблицу месяцев")
    arow = _fin_find(ag, name, afirst, aitog)
    ashift = 0
    if arow is None:
        _fin_insert_rows(book, aws, aitog, 1)
        arow = aitog
        ashift = 1
    r1 = arow + 1
    ti = aitog + ashift + 1
    ups = [
        {"range": f"A{r1}", "values": [[name.capitalize()]]},
        {"range": f"B{r1}", "values": [[f"='{name}'!D{m_post}"]]},
        {"range": f"C{r1}", "values": [[f"='{name}'!D{m_rash}"]]},
        {"range": f"D{r1}", "values": [[f"=B{r1}-C{r1}"]]},
        {"range": f"E{r1}", "values": [[f"='{name}'!D{m_oper}"]]},
        {"range": "A1", "values": [[
            f'ТОО "Едiл и компания" — анализ, выводы и рекомендации (январь–{name} {year} г.)']]},
    ]
    for cl in "BCDE":
        ups.append({"range": f"{cl}{ti}", "values": [[f"=SUM({cl}{afirst + 1}:{cl}{ti - 1})"]]})
    aws.batch_update(ups, value_input_option="USER_ENTERED")
    # максимумы/минимумы — по уже посчитанным значениям таблицы
    try:
        vals = aws.get(f"A{afirst + 1}:C{ti - 1}")
        rows_data = [(str(r[0]), parse_num(r[1]), parse_num(r[2]))
                     for r in vals if r and len(r) >= 3 and str(r[0]).strip()]
        h2 = _fin_find(ag, "2. максимумы")
        if rows_data and h2 is not None:
            mx_p = max(rows_data, key=lambda x: x[1]); mn_p = min(rows_data, key=lambda x: x[1])
            mx_r = max(rows_data, key=lambda x: x[2]); mn_r = min(rows_data, key=lambda x: x[2])
            t1 = (f"• Поступления: максимум — {mx_p[0].upper()} ({_mln(mx_p[1])} млн), "
                  f"минимум — {mn_p[0].upper()} ({_mln(mn_p[1])} млн).")
            t2 = (f"• Расходы: максимум — {mx_r[0].upper()} ({_mln(mx_r[1])} млн), "
                  f"минимум — {mn_r[0].upper()} ({_mln(mn_r[1])} млн).")
            aws.batch_update([
                {"range": f"A{h2 + 2 + ashift}", "values": [[t1]]},
                {"range": f"A{h2 + 3 + ashift}", "values": [[t2]]},
            ], value_input_option="USER_ENTERED")
    except Exception as e:
        logger.error(f"analiz max/min error: {e}")
    # структура операционных расходов — из «Выводов» (столбец Итого)
    try:
        h4 = _fin_find(ag, "4. структура")
        if h4 is not None and ri and rf and rl:
            vv = vyv.get(f"{itog}1:{itog}{max(vrows.values())}")
            def gv(row):
                i = row - 1
                return parse_num(vv[i][0]) if i < len(vv) and vv[i] else 0
            tot = gv(ri)
            if tot:
                pr = vrows.get("прочие расходы")
                items = [("Лизинг", gv(rl)), ("ФОТ", gv(rf)),
                         ("Прочие", gv(pr) if pr else 0)]
                line = "; ".join(
                    f"{nm} — {v / tot * 100:.1f}".replace(".", ",") + f"% ({_mln(v)} млн)"
                    for nm, v in items) + "."
                aws.batch_update([
                    {"range": f"A{h4 + 1 + ashift}", "values": [[
                        f"4. СТРУКТУРА ОПЕРАЦИОННЫХ РАСХОДОВ ЗА {n_mon} МЕС ({_mln(tot)} млн)"]]},
                    {"range": f"A{h4 + 2 + ashift}", "values": [[line]]},
                ], value_input_option="USER_ENTERED")
    except Exception as e:
        logger.error(f"analiz structure error: {e}")


# ---------- авто-обновление «Калькуляции» и финотчёта после каждой записи ----------
_refresh_lock = asyncio.Lock()
_refresh_again = False


async def _run_refresh():
    """Обновляет «Калькуляцию» и финотчёт. Параллельные вызовы совмещает
    (замок исключает гонки при вставке строк во вкладку финотчёта)."""
    global _refresh_again
    if _refresh_lock.locked():
        _refresh_again = True
        return
    async with _refresh_lock:
        while True:
            _refresh_again = False
            month = datetime.now().month
            try:
                await asyncio.to_thread(_fill_kalk, month)
            except Exception as e:
                logger.error(f"auto-kalk error: {e}")
            try:
                await asyncio.to_thread(_fill_fin, month)
            except Exception as e:
                logger.error(f"auto-fin error: {e}")
            try:
                await asyncio.to_thread(_fill_reports, month)
            except Exception as e:
                logger.error(f"auto-reports error: {e}")
            if not _refresh_again:
                break


def schedule_refresh():
    """Фоновое обновление через несколько секунд после записи (ответ бота не задерживает)."""
    async def _job():
        await asyncio.sleep(5)
        await _run_refresh()
    try:
        asyncio.get_running_loop().create_task(_job())
    except RuntimeError:
        pass  # нет цикла событий (например, в тестах) — пропускаем


def main():
    app = Application.builder().token(BOT_TOKEN).post_init(_post_init).build()

    prod_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^✍️ Ввод производства$"), manual_prod_start)],
        states={
            P_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_prod_date)],
            P_FIO: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_prod_fio)],
            P_TIRES: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_prod_tires)],
            P_THREAD: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_prod_thread)],
            P_F01: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_prod_f01)],
            P_F12: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_prod_f12)],
            P_F24: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_prod_f24)],
            P_F46: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_prod_f46)],
            P_F68: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_prod_f68)],
            P_CORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_prod_cord)],
            P_NOTE: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_prod_note)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    sale_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^✍️ Ввод реализации$"), manual_sale_start)],
        states={
            S_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_sale_date)],
            S_BUYER: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_sale_buyer)],
            S_PAYTYPE: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_sale_paytype)],
            S_BANK: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_sale_bank)],
            S_FRAC: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_sale_frac)],
            S_KG: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_sale_kg)],
            S_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_sale_price)],
            S_SUM_VAT: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_sale_sum_vat)],
            S_VAT: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_sale_vat)],
            S_NOTE: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_sale_note)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    # Общие шаги ввода расхода — используются и из «💸 Ввод расхода»,
    # и когда фото расходного документа пришло без старта (через photo_conv).
    def _exp_states():
        return {
            E_DATE: [MessageHandler(filters.PHOTO, manual_exp_photo), MessageHandler(filters.TEXT & ~filters.COMMAND, manual_exp_date)],
            E_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_exp_amount)],
            E_GROUP: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_exp_group)],
            E_CATEGORY: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_exp_category)],
            E_SOURCE: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_exp_source)],
            E_CONTRAGENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_exp_contragent)],
            E_NOTE: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_exp_note)],
            E_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_exp_confirm)],
            E_SUP_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, exp_sup_confirm)],
            E_SUP_NEW: [MessageHandler(filters.TEXT & ~filters.COMMAND, exp_sup_new)],
            E_SUP_EDIT: [MessageHandler(filters.TEXT & ~filters.COMMAND, exp_sup_edit)],
        }

    # NEW: диалог фото-производства с подтверждением дубля.
    # entry_point — приход фото; если дубль, переходим в PHOTO_PROD_CONFIRM и ждём Да/Нет.
    photo_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.PHOTO, handle_photo)],
        states={
            PHOTO_TYPE_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, photo_type_confirm)],
            PHOTO_PROD_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, photo_prod_confirm)],
            PHOTO_SALE_PAY: [MessageHandler(filters.TEXT & ~filters.COMMAND, photo_sale_paytype)],
            PHOTO_SALE_BANK: [MessageHandler(filters.TEXT & ~filters.COMMAND, photo_sale_bank)],
            PHOTO_SALE_KG: [MessageHandler(filters.TEXT & ~filters.COMMAND, photo_sale_kg)],
            PHOTO_SALE_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, photo_sale_price)],
            PHOTO_SALE_VAT: [MessageHandler(filters.TEXT & ~filters.COMMAND, photo_sale_vat)],
            PAYROLL_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, payroll_date)],
            PAYROLL_BANK: [MessageHandler(filters.TEXT & ~filters.COMMAND, payroll_bank)],
            PAYROLL_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, payroll_confirm)],
            TABEL_MONTH: [MessageHandler(filters.TEXT & ~filters.COMMAND, tabel_month)],
            TABEL_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, tabel_confirm)],
            **_exp_states(),  # фото-расход продолжает шаги расхода прямо здесь
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("ostatok", ostatok))
    app.add_handler(CommandHandler("last", last_records))
    app.add_handler(CommandHandler("cancel", cancel))
    app.add_handler(CommandHandler("backup", cmd_backup))
    app.add_handler(CommandHandler("svod", cmd_svod))
    app.add_handler(CommandHandler("kalk", cmd_kalk))
    app.add_handler(CommandHandler("fin", cmd_fin))
    exp_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^💸 Ввод расхода$"), manual_exp_start)],
        states=_exp_states(),
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    # «Спасательный круг»: в каждом шаге каждого диалога кнопки меню и «❌ Отмена»
    # обрабатываются первыми — незаконченный ввод сбрасывается, бот не залипает.
    _escape = MessageHandler(filters.Regex(MENU_ESCAPE_RE), conv_menu_escape)
    for _conv in (prod_conv, sale_conv, exp_conv, photo_conv):
        for _handlers in _conv.states.values():
            _handlers.insert(0, _escape)

    app.add_handler(prod_conv)
    app.add_handler(sale_conv)
    app.add_handler(exp_conv)
    app.add_handler(photo_conv)  # после exp_conv: фото в режиме «Ввод расхода» ловит exp_conv
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    if WEBHOOK_URL:
        logger.info(f"Starting webhook on port {PORT}")
        app.run_webhook(
            listen="0.0.0.0",
            port=PORT,
            webhook_url=f"{WEBHOOK_URL}/{BOT_TOKEN}",
            url_path=BOT_TOKEN,
            drop_pending_updates=True
        )
    else:
        logger.info("Starting polling...")
        app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
